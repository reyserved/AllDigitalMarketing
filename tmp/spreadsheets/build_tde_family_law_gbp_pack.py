#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import subprocess
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/Applications/Antigravity/ROCKET CLICKS/ALL DIGITAL MARKETING')
INPUT_CSV = ROOT / 'Copy of TDE _ Site Audit & Content Mapping _ 1.2026 - Sheet30.csv'
OUT_DIR = ROOT / 'output' / 'spreadsheet'
OUT_XLSX = OUT_DIR / 'TDE_Family_Law_GBP_Posting_Calendar_Mar-Dec_2026.xlsx'
OUT_SUB2_CSV = OUT_DIR / 'tde-family-law-gbp-supporting-topics.csv'
OUT_QA = OUT_DIR / 'tde-family-law-topics-trends-qa-2026.txt'
EVIDENCE_DIR = OUT_DIR / 'tde-family-law-trends-evidence-2026'

START_DATE = date(2026, 3, 1)
END_DATE = date(2026, 12, 27)
WEEK_COUNT = 44


@dataclass(frozen=True)
class TopicBlueprint:
    service_category: str
    subservice_category: Optional[str]
    supporting_topic: str
    topic_target: str
    cta_button_label: str


BLUEPRINT_BY_URL: Dict[str, TopicBlueprint] = {
    'https://tdefamilylaw.com/family-law/divorce/': TopicBlueprint(
        service_category='Divorce',
        subservice_category=None,
        supporting_topic='Georgia divorce filing steps and timeline checkpoints',
        topic_target='filing sequence deadlines and procedural milestones',
        cta_button_label='Schedule consult',
    ),
    'https://tdefamilylaw.com/family-law/uncontested-divorce-lawyer/': TopicBlueprint(
        service_category='Divorce',
        subservice_category='Uncontested Divorce',
        supporting_topic='When uncontested divorce is realistic in Georgia',
        topic_target='eligibility conditions and agreement readiness checks',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/contested-divorce-lawyer/': TopicBlueprint(
        service_category='Divorce',
        subservice_category='Contested Divorce',
        supporting_topic='What makes a Georgia divorce contested and how to prepare',
        topic_target='dispute triggers evidence planning and litigation readiness',
        cta_button_label='Start now',
    ),
    'https://tdefamilylaw.com/family-law/high-asset-divorce/': TopicBlueprint(
        service_category='Divorce',
        subservice_category='High Net Worth Divorce',
        supporting_topic='Asset tracing priorities in Georgia high net worth divorce',
        topic_target='valuation discovery and separate-vs-marital asset analysis',
        cta_button_label='Schedule consult',
    ),
    'https://tdefamilylaw.com/family-law/lgbtq-divorce/': TopicBlueprint(
        service_category='Divorce',
        subservice_category='LGBTQ+ Divorce',
        supporting_topic='Legal planning issues in LGBTQ+ divorce matters in Georgia',
        topic_target='parenting property and procedural considerations for LGBTQ+ spouses',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/legal-separation/': TopicBlueprint(
        service_category='Divorce',
        subservice_category='Legal Separation',
        supporting_topic='When legal separation is a better fit in Georgia family law',
        topic_target='separation-vs-divorce decision factors and legal consequences',
        cta_button_label='Learn more',
    ),
    'https://tdefamilylaw.com/family-law/child-custody-lawyer/': TopicBlueprint(
        service_category='Child Custody',
        subservice_category=None,
        supporting_topic='Best-interest factors used in Georgia custody decisions',
        topic_target='child-focused judicial criteria and evidentiary priorities',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/custody-modification-lawyer/': TopicBlueprint(
        service_category='Child Custody',
        subservice_category='Custody Modification',
        supporting_topic='When Georgia custody orders qualify for modification',
        topic_target='material-change threshold and modification filing requirements',
        cta_button_label='Start now',
    ),
    'https://tdefamilylaw.com/family-law/parental-visitation/': TopicBlueprint(
        service_category='Child Custody',
        subservice_category='Parental Visitation',
        supporting_topic='How Georgia visitation schedules are structured',
        topic_target='parenting-time framework holiday allocation and logistics',
        cta_button_label='Learn more',
    ),
    'https://tdefamilylaw.com/family-law/visitation-modification/': TopicBlueprint(
        service_category='Child Custody',
        subservice_category='Visitation Modification',
        supporting_topic='Evidence needed to modify visitation in Georgia',
        topic_target='proof standards for parenting-time schedule changes',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/alimony-spousal-support/': TopicBlueprint(
        service_category='Spousal Support',
        subservice_category=None,
        supporting_topic='How Georgia courts assess alimony amount and duration',
        topic_target='financial need ability-to-pay and duration factors',
        cta_button_label='Schedule consult',
    ),
    'https://tdefamilylaw.com/family-law/prenuptial-agreement-atlanta/': TopicBlueprint(
        service_category='Prenuptial Agreements',
        subservice_category=None,
        supporting_topic='What makes a prenup enforceable in Georgia',
        topic_target='contract validity standards disclosure and execution rules',
        cta_button_label='Learn more',
    ),
    'https://tdefamilylaw.com/family-law/postnuptial-agreement-lawyer/': TopicBlueprint(
        service_category='Postnuptial Agreements',
        subservice_category=None,
        supporting_topic='When a Georgia postnuptial agreement should be used',
        topic_target='marriage-stage risk planning and enforceability requirements',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/property-division-enforcement/': TopicBlueprint(
        service_category='Property Division',
        subservice_category=None,
        supporting_topic='Equitable division fundamentals for Georgia marital property',
        topic_target='classification valuation and distribution principles',
        cta_button_label='Start now',
    ),
    'https://tdefamilylaw.com/family-law/parental-rights/': TopicBlueprint(
        service_category='Parental Rights',
        subservice_category=None,
        supporting_topic='How parental rights are established and defended in Georgia',
        topic_target='rights allocation parental fitness and procedural safeguards',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/fathers-rights-lawyer/': TopicBlueprint(
        service_category='Parental Rights',
        subservice_category='Fathers Rights',
        supporting_topic='Building a fathers’ rights evidence record in Georgia',
        topic_target='documentation strategy for parenting role and decision-making participation',
        cta_button_label='Schedule consult',
    ),
    'https://tdefamilylaw.com/family-law/mothers-rights-lawyer-in-atlanta/': TopicBlueprint(
        service_category='Parental Rights',
        subservice_category='Mothers Rights',
        supporting_topic='Protecting mothers’ rights in Georgia family law proceedings',
        topic_target='custody support and parental-decision protections',
        cta_button_label='Contact us',
    ),
    'https://tdefamilylaw.com/family-law/grandparents-rights-lawyer/': TopicBlueprint(
        service_category='Parental Rights',
        subservice_category='Grandparents Rights',
        supporting_topic='When grandparents can seek visitation or custody in Georgia',
        topic_target='standing requirements and best-interest evidentiary burden',
        cta_button_label='Learn more',
    ),
    'https://tdefamilylaw.com/family-law/paternity-lawyer-in-atlanta/': TopicBlueprint(
        service_category='Paternity',
        subservice_category=None,
        supporting_topic='Steps to establish legal paternity in Georgia',
        topic_target='acknowledgment testing and court-order pathways',
        cta_button_label='Start now',
    ),
}

CATEGORY_CONTEXT = {
    'Divorce': 'In Georgia divorce matters, timelines, financial disclosures, and enforceable court orders usually drive outcomes more than reactive decisions.',
    'Uncontested Divorce': 'In Georgia uncontested divorce matters, courts expect complete agreement terms on property, support, and parenting before streamlined resolution is realistic.',
    'Contested Divorce': 'In contested Georgia divorce cases, preparation around disputed facts, records, and deadlines often determines leverage and court readiness.',
    'High Net Worth Divorce': 'In high net worth Georgia divorce matters, accurate asset tracing, valuation support, and tax-aware planning are central to fair outcomes.',
    'LGBTQ+ Divorce': 'In LGBTQ+ Georgia divorce matters, parenting history, legal parentage, and property documentation can significantly affect final outcomes.',
    'Legal Separation': 'In Georgia legal separation matters, temporary arrangements for finances and parenting should be structured carefully because they shape later proceedings.',
    'Child Custody': "In Georgia child custody matters, the court's best-interest standard focuses on stability, safety, and each parent's ability to meet a child's needs.",
    'Custody Modification': 'In Georgia custody modification cases, a substantial and continuing change must be documented before the court revisits prior orders.',
    'Parental Visitation': 'In Georgia visitation planning, clear schedules, exchange logistics, and holiday rules reduce conflict and improve long-term compliance.',
    'Visitation Modification': 'In Georgia visitation modification cases, objective records and child-centered reasons are critical when seeking schedule changes.',
    'Spousal Support': 'In Georgia spousal support matters, need, ability to pay, and duration evidence are usually central to enforceable support outcomes.',
    'Prenuptial Agreements': 'In Georgia prenup matters, enforceability depends on proper execution, transparent disclosures, and terms that can withstand legal challenge.',
    'Postnuptial Agreements': 'In Georgia postnuptial planning, valid agreements require careful drafting, full financial clarity, and procedural fairness.',
    'Property Division': 'In Georgia property division matters, classifying, valuing, and allocating assets with documentation is essential to equitable distribution.',
    'Parental Rights': 'In Georgia parental rights matters, legal standing, factual consistency, and child-focused evidence shape how rights are protected or challenged.',
    'Fathers Rights': "In Georgia fathers' rights matters, consistent involvement records and objective evidence help protect parenting time and decision-making authority.",
    'Mothers Rights': "In Georgia mothers' rights matters, clear documentation of caregiving, stability, and child needs supports strong legal positioning.",
    'Grandparents Rights': 'In Georgia grandparents rights matters, standing and best-interest proof requirements are specific and must be met with focused evidence.',
    'Paternity': 'In Georgia paternity matters, choosing the right acknowledgment or testing pathway affects custody, support, and parental rights outcomes.',
}

OPENERS = [
    'Families often reach this issue when uncertainty is high and the next legal step is not obvious.',
    'This topic usually becomes urgent once a case starts affecting daily routines, finances, or parenting decisions.',
    'Many people feel stuck at this stage because the rules are technical and the consequences are long-term.',
    'When this question appears in a case, clear guidance can prevent avoidable conflict and costly delays.',
    'This is one of the most important planning points in Georgia family law because early choices shape final outcomes.',
    'People often ask about this after receiving new information, and timing matters more than most expect.',
    'If this issue is in your case, structured preparation can reduce stress and improve decision quality.',
]

REASSURANCE_LINES = [
    'You can move forward with a calmer, step-by-step plan when expectations and documents are organized early.',
    'With the right preparation, most families can reduce pressure and make decisions with more confidence.',
    'A focused legal strategy helps turn uncertainty into clear options and practical next actions.',
    'Careful planning now usually prevents avoidable setbacks and keeps your case moving in a stable direction.',
]


def slug_title(url: str) -> str:
    parts = [p for p in url.rstrip('/').split('/') if p]
    if not parts:
        return 'Family Law'
    last = parts[-1]
    return ' '.join(w.capitalize() for w in re.split(r'[-_]', last))


def normalize_keyword(topic: str) -> str:
    lowered = topic.lower().strip()
    if re.search(r'\bgeorgia\b', lowered):
        return lowered
    return f"{lowered} georgia"


def lens_name(service: str, subservice: Optional[str]) -> str:
    return subservice if subservice else service


def generate_copy(
    title: str,
    topic_target: str,
    service_category: str,
    subservice_category: Optional[str],
    page_label: str,
    week_index: int,
) -> str:
    lens = lens_name(service_category, subservice_category)
    opener = OPENERS[week_index % len(OPENERS)]
    context = CATEGORY_CONTEXT[lens]
    reassurance = REASSURANCE_LINES[week_index % len(REASSURANCE_LINES)]
    sentence_1 = (
        f"{title} is a critical issue for many Georgia families because legal details in this area can shape long-term parenting, financial, and case outcomes."
    )
    sentence_2 = f"{opener.rstrip('.')} {context}"
    sentence_3 = (
        f"This post focuses on {topic_target}, and {reassurance.lower()}"
    )
    sentence_4 = (
        f"Review our {page_label} page and contact TDE Family Law to schedule a consultation about the next step in your Georgia matter."
    )
    return ' '.join([sentence_1, sentence_2, sentence_3, sentence_4])


def word_count(text: str) -> int:
    return len(re.findall(r"\b[\w'’-]+\b", text))


def sentence_count(text: str) -> int:
    return len([s for s in re.split(r'(?<=[.!?])\s+', text.strip()) if s.strip()])


def apply_table_style(ws, header_row: int, start_col: int, end_col: int, freeze_cell: str):
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(end_col)}{ws.max_row}"


def write_probe(evidence_dir: Path) -> List[dict]:
    evidence_dir.mkdir(parents=True, exist_ok=True)
    probes = [
        {
            'name': 'trends_home',
            'url': 'https://trends.google.com/trends/',
        },
        {
            'name': 'trends_explore_probe',
            'url': 'https://trends.google.com/trends/api/explore?hl=en-US&tz=360&req=%7B%22comparisonItem%22:%5B%7B%22keyword%22:%22divorce%20in%20georgia%22,%22geo%22:%22US-GA%22,%22time%22:%222023-01-01%202026-02-20%22%7D%5D,%22category%22:0,%22property%22:%22%22%7D',
        },
        {
            'name': 'trending_rss_ga',
            'url': 'https://trends.google.com/trending/rss?geo=US-GA',
        },
    ]
    results = []
    for p in probes:
        body_file = evidence_dir / f"{p['name']}.raw"
        cmd = [
            'curl',
            '-sS',
            '--max-time',
            '30',
            '-D',
            '-',
            p['url'],
        ]
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
            raw = proc.stdout + '\n' + proc.stderr
            body_file.write_text(raw, encoding='utf-8')
            status_match = re.search(r'HTTP/\d(?:\.\d)?\s+(\d{3})', proc.stdout)
            status = int(status_match.group(1)) if status_match else None
            results.append({
                'probe': p['name'],
                'url': p['url'],
                'status': status,
                'exit_code': proc.returncode,
                'file': str(body_file),
            })
        except Exception as exc:
            body_file.write_text(str(exc), encoding='utf-8')
            results.append({
                'probe': p['name'],
                'url': p['url'],
                'status': None,
                'exit_code': -1,
                'file': str(body_file),
            })

    (evidence_dir / 'access_probe_summary.json').write_text(json.dumps(results, indent=2), encoding='utf-8')
    return results


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    mapped_rows: List[dict] = []
    with INPUT_CSV.open(newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            mapped_rows.append(row)

    if len(mapped_rows) != 19:
        raise ValueError(f'Expected 19 mapped rows, got {len(mapped_rows)}')

    # Build base aligned records in source order.
    base = []
    for row in mapped_rows:
        url = row['Existing Client URL'].strip()
        if url not in BLUEPRINT_BY_URL:
            raise KeyError(f'No blueprint configured for URL: {url}')
        bp = BLUEPRINT_BY_URL[url]
        base.append({
            'service_type': row['Service Type'].strip(),
            'service_category': bp.service_category,
            'subservice_category': bp.subservice_category,
            'mapped_seed_keyword': row['Target Seed Keyword'].strip(),
            'mapped_page_url': url,
            'supporting_topic': bp.supporting_topic,
            'topic_target': bp.topic_target,
            'primary_keyword_target': normalize_keyword(bp.supporting_topic),
            'search_intent': 'Service + Informational',
            'cta_button_label': bp.cta_button_label,
        })

    # 44-week assignment: each core topic at least twice; six topics three times.
    assignments = list(range(19)) + list(range(19)) + [0, 6, 10, 14, 18, 13]
    if len(assignments) != WEEK_COUNT:
        raise ValueError('Assignment count mismatch')

    # Build Subtask2 (44 rows) + Subtask3 calendar (44 rows)
    subtask2_rows = []
    calendar_rows = []
    date_cursor = START_DATE
    for i, idx in enumerate(assignments, start=1):
        if date_cursor > END_DATE:
            raise ValueError('Date range exceeded')
        topic_id = f'TOPIC-{i:03d}'
        rec = base[idx]
        lens = lens_name(rec['service_category'], rec['subservice_category'])
        note = f'Week {i:02d} | {lens}'

        s2 = {
            'topic_id': topic_id,
            'service_category': rec['service_category'],
            'subservice_category': rec['subservice_category'],
            'mapped_seed_keyword': rec['mapped_seed_keyword'],
            'mapped_page_url': rec['mapped_page_url'],
            'supporting_topic': rec['supporting_topic'],
            'topic_target': rec['topic_target'],
            'primary_keyword_target': rec['primary_keyword_target'],
            'search_intent': rec['search_intent'],
            'cta_button_label': rec['cta_button_label'],
            'notes': note,
        }
        subtask2_rows.append(s2)

        page_label = slug_title(rec['mapped_page_url'])
        copy = generate_copy(
            title=rec['supporting_topic'],
            topic_target=rec['topic_target'],
            service_category=rec['service_category'],
            subservice_category=rec['subservice_category'],
            page_label=page_label,
            week_index=i - 1,
        )

        cdesc = (
            f"Empathetic Georgia-focused summary for {rec['topic_target']}, with clear reassurance and CTA to our {page_label} page."
        )

        calendar_rows.append({
            'publish_date': datetime.combine(date_cursor, datetime.min.time()),
            'topic_id': topic_id,
            'post_title': rec['supporting_topic'],
            'keyword_target': rec['primary_keyword_target'],
            'topic_target': rec['topic_target'],
            'link_destination': rec['mapped_page_url'],
            'content_description': cdesc,
            'post_copy_80_150_words': copy,
            'cta_button_label': rec['cta_button_label'],
            'rotation_bucket': lens,
            'image_thumbnail': None,
            'image_source_url': None,
            'image_alt_text': None,
        })

        date_cursor += timedelta(days=7)

    # Write Subtask2 CSV
    sub2_headers = [
        'topic_id',
        'service_category',
        'subservice_category',
        'mapped_seed_keyword',
        'mapped_page_url',
        'supporting_topic',
        'topic_target',
        'primary_keyword_target',
        'search_intent',
        'cta_button_label',
        'notes',
    ]
    with OUT_SUB2_CSV.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=sub2_headers)
        writer.writeheader()
        writer.writerows(subtask2_rows)

    # Build workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Subtask1_Mapped
    ws1 = wb.active
    ws1.title = 'Subtask1_Mapped'
    ws1.merge_cells('A1:C1')
    ws1['A1'] = 'TDE Family Law GBP Posting Pack - Subtask 1 Mapped'
    ws1['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws1['A2'] = 'Source: Copy of TDE _ Site Audit & Content Mapping _ 1.2026 - Sheet30.csv'
    headers1 = ['Service Type', 'Target Seed Keyword', 'Existing Client URL']
    for c, h in enumerate(headers1, start=1):
        ws1.cell(4, c, h)
    for r_idx, row in enumerate(mapped_rows, start=5):
        ws1.cell(r_idx, 1, row['Service Type'])
        ws1.cell(r_idx, 2, row['Target Seed Keyword'])
        ws1.cell(r_idx, 3, row['Existing Client URL'])
    apply_table_style(ws1, header_row=4, start_col=1, end_col=3, freeze_cell='A5')
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 38
    ws1.column_dimensions['C'].width = 65

    # Sheet 2: Subtask2_Topics
    ws2 = wb.create_sheet('Subtask2_Topics')
    ws2.merge_cells('A1:K1')
    ws2['A1'] = 'TDE Family Law GBP Posting Pack - Subtask 2 Supporting Topics (AZ-style fallback)'
    ws2['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws2['A2'] = 'No reliable Google Trends pull in this runtime; strict topical alignment used.'

    for c, h in enumerate(sub2_headers, start=1):
        ws2.cell(4, c, h)
    for r_idx, row in enumerate(subtask2_rows, start=5):
        for c, h in enumerate(sub2_headers, start=1):
            ws2.cell(r_idx, c, row[h])

    apply_table_style(ws2, header_row=4, start_col=1, end_col=11, freeze_cell='A5')
    widths2 = [12, 19, 21, 36, 52, 56, 44, 46, 22, 16, 28]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(5, ws2.max_row + 1):
        for c in range(1, 12):
            ws2.cell(r, c).alignment = Alignment(vertical='top', wrap_text=True)

    # Sheet 3: Subtask3_Calendar
    ws3 = wb.create_sheet('Subtask3_Calendar')
    ws3.merge_cells('A1:M1')
    ws3['A1'] = 'TDE Family Law GBP Posting Calendar - March to December 2026 (AZ-style fallback)'
    ws3['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws3['A2'] = '44 weekly rows with strict topic/title/copy alignment. Image fields intentionally left blank.'

    headers3 = [
        'publish_date',
        'topic_id',
        'post_title',
        'keyword_target',
        'topic_target',
        'link_destination',
        'content_description',
        'post_copy_80_150_words',
        'cta_button_label',
        'rotation_bucket',
        'image_thumbnail',
        'image_source_url',
        'image_alt_text',
    ]
    for c, h in enumerate(headers3, start=1):
        ws3.cell(4, c, h)

    for r_idx, row in enumerate(calendar_rows, start=5):
        for c, h in enumerate(headers3, start=1):
            ws3.cell(r_idx, c, row[h])

    apply_table_style(ws3, header_row=4, start_col=1, end_col=13, freeze_cell='A5')
    widths3 = [14, 11, 45, 42, 38, 52, 60, 95, 16, 24, 18, 24, 28]
    for i, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    for r in range(5, ws3.max_row + 1):
        ws3.row_dimensions[r].height = 90
        for c in range(1, 14):
            ws3.cell(r, c).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(r, 1).number_format = 'yyyy-mm-dd'

    # Sheet 4: Drive Image Links (placeholder structure)
    ws4 = wb.create_sheet('Drive Image Links')
    ws4.merge_cells('A1:C1')
    ws4['A1'] = 'Drive Image Links (Placeholder - to be populated later)'
    ws4['A1'].font = Font(size=13, bold=True, color='1F4E78')
    h4 = ['topic_id', 'image_thumbnail', 'google_drive_image_link']
    for c, h in enumerate(h4, start=1):
        ws4.cell(4, c, h)
    for i in range(1, WEEK_COUNT + 1):
        ws4.cell(4 + i, 1, f'TOPIC-{i:03d}')
        ws4.cell(4 + i, 2, None)
        ws4.cell(4 + i, 3, None)
    apply_table_style(ws4, header_row=4, start_col=1, end_col=3, freeze_cell='A5')
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 24
    ws4.column_dimensions['C'].width = 55

    wb.save(OUT_XLSX)

    # Trends access probe artifacts
    probe_results = write_probe(EVIDENCE_DIR)

    # QA checks
    errors: List[str] = []

    if len(subtask2_rows) != WEEK_COUNT:
        errors.append(f'Subtask2 row count expected {WEEK_COUNT}, got {len(subtask2_rows)}')
    if len(calendar_rows) != WEEK_COUNT:
        errors.append(f'Subtask3 row count expected {WEEK_COUNT}, got {len(calendar_rows)}')

    mapped_urls = {r['Existing Client URL'].strip() for r in mapped_rows}
    for r in subtask2_rows:
        if r['mapped_page_url'] not in mapped_urls:
            errors.append(f"Unmapped URL in Subtask2: {r['mapped_page_url']}")

    idx2 = {r['topic_id']: r for r in subtask2_rows}
    for r in calendar_rows:
        t = idx2[r['topic_id']]
        if r['post_title'] != t['supporting_topic']:
            errors.append(f"Title mismatch for {r['topic_id']}")
        if r['topic_target'] != t['topic_target']:
            errors.append(f"Topic target mismatch for {r['topic_id']}")
        if r['keyword_target'] != t['primary_keyword_target']:
            errors.append(f"Keyword mismatch for {r['topic_id']}")
        wc = word_count(r['post_copy_80_150_words'])
        sc = sentence_count(r['post_copy_80_150_words'])
        if not (80 <= wc <= 150):
            errors.append(f"Word count out of range for {r['topic_id']}: {wc}")
        if not (2 <= sc <= 4):
            errors.append(f"Sentence count out of range for {r['topic_id']}: {sc}")
        if not re.search(r'\b(review|contact|schedule|consultation|call|visit)\b', r['post_copy_80_150_words'], re.I):
            errors.append(f"CTA not detected for {r['topic_id']}")

    # Strict topical alignment token checks
    disallowed = {
        'Divorce': ['custody modification', 'visitation modification', 'paternity'],
        'Child Custody': ['alimony', 'spousal support', 'prenup', 'postnuptial'],
        'Spousal Support': ['custody modification', 'visitation modification', 'paternity'],
        'Parental Rights': ['alimony', 'prenup', 'postnuptial'],
        'Paternity': ['alimony', 'prenup', 'postnuptial'],
    }
    for r in subtask2_rows:
        lens = lens_name(r['service_category'], r['subservice_category'])
        text = f"{r['supporting_topic']} {r['topic_target']}".lower()
        base_lens = r['service_category']
        for bad in disallowed.get(base_lens, []):
            if bad in text:
                errors.append(f"Potential topical drift in {r['topic_id']} ({lens}): found '{bad}'")

    with OUT_QA.open('w', encoding='utf-8') as f:
        f.write('TDE Family Law GBP Pack QA\n')
        f.write(f'Generated at: {datetime.utcnow().isoformat()}Z\n\n')
        f.write('Google Trends Access Probes:\n')
        for pr in probe_results:
            f.write(
                f"- {pr['probe']}: status={pr['status']} exit_code={pr['exit_code']} file={pr['file']}\n"
            )
        probe_ok = any((pr.get('status') == 200) for pr in probe_results)
        if not probe_ok:
            f.write('\nConclusion: Reliable Google Trends API/endpoint pulls were not available in this runtime. Applied AZ-style fallback workflow for TDE.\n')
        else:
            f.write('\nConclusion: At least one Trends endpoint responded successfully.\n')

        f.write('\nCounts:\n')
        f.write(f'- Input mapped rows: {len(mapped_rows)}\n')
        f.write(f'- Subtask2 rows: {len(subtask2_rows)}\n')
        f.write(f'- Subtask3 rows: {len(calendar_rows)}\n')

        reuse = {}
        for r in subtask2_rows:
            key = (r['supporting_topic'], r['mapped_page_url'])
            reuse[key] = reuse.get(key, 0) + 1
        f.write('\nTopic Reuse Distribution:\n')
        counts = sorted(reuse.values())
        f.write(f'- min reuse: {min(counts)}\n')
        f.write(f'- max reuse: {max(counts)}\n')

        f.write('\nValidation:\n')
        if errors:
            f.write(f'- FAIL ({len(errors)} issues)\n')
            for err in errors:
                f.write(f'  - {err}\n')
        else:
            f.write('- PASS (all checks)\n')

    print(f'Wrote workbook: {OUT_XLSX}')
    print(f'Wrote subtask2 csv: {OUT_SUB2_CSV}')
    print(f'Wrote QA report: {OUT_QA}')
    print(f'Wrote evidence folder: {EVIDENCE_DIR}')


if __name__ == '__main__':
    main()
