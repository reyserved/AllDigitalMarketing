from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


SOURCE_CSV = Path(
    "/Applications/Antigravity/ROCKET CLICKS/ALL DIGITAL MARKETING/02-audits/content/Omnipress_OnPage_Opportunities_Pages_Weighted.csv"
)
MAPPING_CLUSTER_CSV = Path(
    "/Applications/Antigravity/ROCKET CLICKS/ALL DIGITAL MARKETING/02-audits/content/Omnipress _ Keyword Research 2.2026 - Keyword Mapping - pages.csv"
)
OUTPUT_XLSX = Path(
    "/Applications/Antigravity/ROCKET CLICKS/ALL DIGITAL MARKETING/02-audits/content/Omnipress_OnPage_Opportunities_Pages_Simplified.xlsx"
)

SIMPLIFIED_HEADERS = [
    "URL",
    "Assigned Topic Cluster",
    "Priority",
    "Priority Reason (Plain English)",
    "Primary Opportunity Keyword",
    "Supporting Opportunity Keywords",
    "Primary Issue",
    "Secondary Issue",
    "Recommended Next Action",
    "Performance Snapshot",
    "Keyword Coverage Snapshot",
    "Metadata Check",
    "Data Confidence",
]

THEME_MAP = {
    "Title/H1 Keyword Alignment": "Primary keyword is not clear in title/H1",
    "Content Depth & Intent Coverage": "Page content does not fully answer search intent",
    "Snippet/CTR Refinement": "Search snippet likely underperforming CTR",
    "Demand Recovery": "Search demand/visibility is declining",
    "Conversion Path Optimization": "Page is under-converting",
    "": "",
}


def clean(value: str | None) -> str:
    return (value or "").strip()


def normalize_url(url: str | None) -> str:
    u = clean(url).replace("http://", "https://")
    if u.endswith("/") and u != "https://omnipress.com/":
        u = u[:-1]
    return u


def build_cluster_lookup() -> dict[str, str]:
    counts: dict[str, Counter[str]] = {}
    with MAPPING_CLUSTER_CSV.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = normalize_url(row.get("Current URL", ""))
            cluster = clean(row.get("Topic Cluster", ""))
            if not url or not cluster:
                continue
            counts.setdefault(url, Counter())[cluster] += 1

    lookup: dict[str, str] = {}
    for url, counter in counts.items():
        lookup[url] = sorted(counter.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
    return lookup


def parse_pct(text: str) -> float | None:
    text = clean(text)
    if not text:
        return None
    text = text.replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def pct_raw(text: str) -> str:
    return clean(text)


def as_int(text: str) -> int:
    text = clean(text)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def as_float(text: str) -> float:
    text = clean(text)
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def simplify_action(action: str) -> str:
    action = clean(action)
    if not action:
        return ""

    sentences = [s.strip() for s in re.split(r"(?<=\.)\s+", action) if s.strip()]
    if not sentences:
        return action

    keyword_sentence_index = None
    for i, sentence in enumerate(sentences):
        if sentence.startswith("Add '") or sentence.startswith("Reinforce '"):
            keyword_sentence_index = i
            break

    selected: list[str] = []
    if keyword_sentence_index is not None:
        selected.append(sentences[keyword_sentence_index])
        for i, sentence in enumerate(sentences):
            if i == keyword_sentence_index:
                continue
            selected.append(sentence)
            if len(selected) == 2:
                break
    else:
        selected = sentences[:2]

    merged = " ".join(selected).strip()
    return merged


def translate_theme(theme: str) -> str:
    return THEME_MAP.get(clean(theme), clean(theme))


def build_performance_snapshot(row: dict[str, str]) -> str:
    return (
        f"L3M C/I/Conv: {pct_raw(row.get('L3M Δ Clicks %', ''))}/"
        f"{pct_raw(row.get('L3M Δ Impr %', ''))}/"
        f"{pct_raw(row.get('L3M Δ Conv %', ''))} | "
        f"YoY C/I/Conv: {pct_raw(row.get('YoY Δ Clicks %', ''))}/"
        f"{pct_raw(row.get('YoY Δ Impr %', ''))}/"
        f"{pct_raw(row.get('YoY Δ Conv %', ''))}"
    )


def build_keyword_snapshot(row: dict[str, str]) -> str:
    partial = as_int(row.get("Partial Count", ""))
    total = as_int(row.get("Keyword Rows (Count)", ""))
    exact = as_int(row.get("Exact Count", ""))
    optimize = as_int(row.get("Optimize Count", ""))
    return f"Partial {partial} of {total} | Exact {exact} | Optimize {optimize}"


def build_metadata_check(row: dict[str, str]) -> str:
    gaps = clean(row.get("Metadata Gaps", ""))
    if not gaps:
        return "No critical metadata gap detected"
    return f"Needs metadata fixes: {gaps}"


def build_confidence(row: dict[str, str]) -> str:
    conf = clean(row.get("Data Confidence", ""))
    count = as_int(row.get("Metrics Available (out of 10)", ""))
    return f"{conf} ({count}/10 metrics)" if conf else f"Unknown ({count}/10 metrics)"


def theme_key(theme: str) -> str:
    t = clean(theme)
    if t == "Conversion Path Optimization":
        return "conversion"
    if t == "Demand Recovery":
        return "demand"
    if t == "Snippet/CTR Refinement":
        return "ctr"
    if t == "Content Depth & Intent Coverage":
        return "content"
    if t == "Title/H1 Keyword Alignment":
        return "title"
    return "other"


def build_priority_reason(row: dict[str, str]) -> str:
    theme = theme_key(row.get("Primary Opportunity Theme", ""))

    l3m_clicks = clean(row.get("L3M Δ Clicks %", ""))
    l3m_impr = clean(row.get("L3M Δ Impr %", ""))
    l3m_conv = clean(row.get("L3M Δ Conv %", ""))
    yoy_clicks = clean(row.get("YoY Δ Clicks %", ""))
    yoy_impr = clean(row.get("YoY Δ Impr %", ""))
    yoy_conv = clean(row.get("YoY Δ Conv %", ""))

    if theme == "conversion":
        conv_parts = []
        if l3m_conv:
            conv_parts.append(f"L3M conv {l3m_conv}")
        if yoy_conv:
            conv_parts.append(f"YoY conv {yoy_conv}")
        base = "; ".join(conv_parts) if conv_parts else "conversion trend is negative"
        return f"Priority driven by conversion decline: {base}."

    if theme == "demand":
        l3m = l3m_impr if l3m_impr else "NA"
        yoy = yoy_impr if yoy_impr else "NA"
        return f"Priority driven by demand loss: impressions fell in L3M ({l3m}) and YoY ({yoy})."

    if theme == "ctr":
        c_l3m = parse_pct(l3m_clicks)
        i_l3m = parse_pct(l3m_impr)
        c_yoy = parse_pct(yoy_clicks)
        i_yoy = parse_pct(yoy_impr)

        if c_l3m is not None and c_l3m < 0 and i_l3m is not None and i_l3m >= 0:
            return f"Priority driven by CTR friction: L3M clicks {l3m_clicks} while L3M impressions {l3m_impr}."
        if c_yoy is not None and c_yoy < 0 and i_yoy is not None and i_yoy >= 0:
            return f"Priority driven by CTR friction: YoY clicks {yoy_clicks} while YoY impressions {yoy_impr}."
        return "Priority driven by CTR refinement signal from clicks vs impressions trend."

    if theme == "content":
        pr = as_float(row.get("Partial Ratio", ""))
        return f"Priority driven by intent coverage gap: partial keyword ratio is {pr:.2f}."

    if theme == "title":
        gaps = clean(row.get("Metadata Gaps", ""))
        if gaps:
            return f"Priority driven by keyword alignment gap in title/H1: {gaps}."
        return "Priority driven by keyword alignment gap in title/H1."

    return "Priority follows weighted model signals across performance, mapping, and metadata."


def supporting_keywords(row: dict[str, str]) -> str:
    vals = [clean(row.get("Top Keyword 2", "")), clean(row.get("Top Keyword 3", ""))]
    vals = [v for v in vals if v]
    if not vals:
        return "None"
    return " | ".join(vals)


def row_to_simplified(row: dict[str, str], cluster_lookup: dict[str, str]) -> dict[str, str]:
    url = clean(row.get("URL", ""))
    assigned_cluster = cluster_lookup.get(normalize_url(url), clean(row.get("Topic Cluster", "")))
    return {
        "URL": url,
        "Assigned Topic Cluster": assigned_cluster,
        "Priority": clean(row.get("Weighted Priority Tier", "")),
        "Priority Reason (Plain English)": build_priority_reason(row),
        "Primary Opportunity Keyword": clean(row.get("Top Keyword 1", "")),
        "Supporting Opportunity Keywords": supporting_keywords(row),
        "Primary Issue": translate_theme(row.get("Primary Opportunity Theme", "")),
        "Secondary Issue": translate_theme(row.get("Secondary Opportunity Theme", "")),
        "Recommended Next Action": simplify_action(row.get("Suggested Action", "")),
        "Performance Snapshot": build_performance_snapshot(row),
        "Keyword Coverage Snapshot": build_keyword_snapshot(row),
        "Metadata Check": build_metadata_check(row),
        "Data Confidence": build_confidence(row),
    }


def glossary_rows() -> list[dict[str, str]]:
    rows = [
        {
            "Term/Metric": "Assigned Topic Cluster",
            "What It Means (Simple)": "Cluster assigned to the page from the keyword mapping sheet.",
            "Why It Matters": "Keeps page grouping consistent with your mapping taxonomy.",
            "How To Use It Quickly": "Filter by cluster to plan related page updates together.",
            "Comprehensive Source Column": "Topic Cluster (from Keyword Mapping - pages.csv)",
        },
        {
            "Term/Metric": "Priority",
            "What It Means (Simple)": "Overall urgency label for this page.",
            "Why It Matters": "Helps teams focus effort where upside/risk is highest.",
            "How To Use It Quickly": "Start with Critical, then High, then Medium.",
            "Comprehensive Source Column": "Weighted Priority Tier",
        },
        {
            "Term/Metric": "Priority Reason (Plain English)",
            "What It Means (Simple)": "One clear sentence explaining why this page is prioritized.",
            "Why It Matters": "Avoids needing to decode technical metrics first.",
            "How To Use It Quickly": "Use it as the summary context for sprint planning.",
            "Comprehensive Source Column": "Derived from Primary Opportunity Theme + performance fields",
        },
        {
            "Term/Metric": "Primary Opportunity Keyword",
            "What It Means (Simple)": "Main keyword to align the page around first.",
            "Why It Matters": "Drives title/H1/body alignment and intent focus.",
            "How To Use It Quickly": "Place in title, H1, intro, and key section headers.",
            "Comprehensive Source Column": "Top Keyword 1",
        },
        {
            "Term/Metric": "Supporting Opportunity Keywords",
            "What It Means (Simple)": "Secondary keywords that support the same page intent.",
            "Why It Matters": "Improves topical breadth and related query coverage.",
            "How To Use It Quickly": "Map each keyword to a dedicated on-page section.",
            "Comprehensive Source Column": "Top Keyword 2, Top Keyword 3",
        },
        {
            "Term/Metric": "Primary Issue",
            "What It Means (Simple)": "Most important on-page problem type for this URL.",
            "Why It Matters": "Guides what to fix first for faster impact.",
            "How To Use It Quickly": "Solve this issue before secondary improvements.",
            "Comprehensive Source Column": "Primary Opportunity Theme",
        },
        {
            "Term/Metric": "Secondary Issue",
            "What It Means (Simple)": "Second-most important issue after the primary one.",
            "Why It Matters": "Helps sequence follow-up improvements.",
            "How To Use It Quickly": "Address after completing the primary issue.",
            "Comprehensive Source Column": "Secondary Opportunity Theme",
        },
        {
            "Term/Metric": "Performance Snapshot",
            "What It Means (Simple)": "Quick view of clicks/impressions/conversions trends in L3M and YoY.",
            "Why It Matters": "Shows if visibility and outcomes are rising or falling.",
            "How To Use It Quickly": "Negative trends suggest recovery/refinement work.",
            "Comprehensive Source Column": "L3M Δ Clicks %, L3M Δ Impr %, L3M Δ Conv %, YoY Δ Clicks %, YoY Δ Impr %, YoY Δ Conv %",
        },
        {
            "Term/Metric": "Keyword Coverage Snapshot",
            "What It Means (Simple)": "How many mapped keywords are partial vs exact and need optimization.",
            "Why It Matters": "Higher partial share usually means weaker intent alignment.",
            "How To Use It Quickly": "Prioritize pages with high partial counts/ratios.",
            "Comprehensive Source Column": "Partial Count, Keyword Rows (Count), Exact Count, Optimize Count",
        },
        {
            "Term/Metric": "Metadata Check",
            "What It Means (Simple)": "Whether title/meta/H1 have critical keyword coverage gaps.",
            "Why It Matters": "Metadata strongly affects relevance and CTR.",
            "How To Use It Quickly": "Fix any listed gaps before deeper content edits.",
            "Comprehensive Source Column": "Metadata Gaps",
        },
        {
            "Term/Metric": "Data Confidence",
            "What It Means (Simple)": "Confidence level based on number of available benchmark metrics.",
            "Why It Matters": "Prevents overconfidence when data is sparse.",
            "How To Use It Quickly": "Treat Low confidence pages as directional, not final.",
            "Comprehensive Source Column": "Data Confidence + Metrics Available (out of 10)",
        },
        {
            "Term/Metric": "Performance Risk Score",
            "What It Means (Simple)": "Risk score from negative L3M/YoY performance deltas.",
            "Why It Matters": "Quantifies trend deterioration severity.",
            "How To Use It Quickly": "Higher score means bigger performance recovery need.",
            "Comprehensive Source Column": "Performance Risk Score",
        },
        {
            "Term/Metric": "Mapping Risk Score",
            "What It Means (Simple)": "Risk score from partial/optimize-heavy keyword mapping mix.",
            "Why It Matters": "Flags pages with weaker keyword-to-page match quality.",
            "How To Use It Quickly": "High score = improve intent coverage and copy targeting.",
            "Comprehensive Source Column": "Mapping Risk Score",
        },
        {
            "Term/Metric": "Metadata Risk Score",
            "What It Means (Simple)": "Risk score from missing or misaligned title/meta/H1 signals.",
            "Why It Matters": "Metadata gaps reduce relevance and snippet performance.",
            "How To Use It Quickly": "Resolve metadata flags to lower this score quickly.",
            "Comprehensive Source Column": "Metadata Risk Score",
        },
        {
            "Term/Metric": "Weighted Priority Score",
            "What It Means (Simple)": "Combined score using performance, mapping, and metadata risks.",
            "Why It Matters": "Creates one comparable urgency score across pages.",
            "How To Use It Quickly": "Sort descending to build execution queue.",
            "Comprehensive Source Column": "Weighted Priority Score",
        },
        {
            "Term/Metric": "Weighted Priority Tier",
            "What It Means (Simple)": "Bucketed urgency from weighted score: Critical/High/Medium/Low.",
            "Why It Matters": "Makes prioritization easier for non-technical stakeholders.",
            "How To Use It Quickly": "Use as planning label in weekly SEO standups.",
            "Comprehensive Source Column": "Weighted Priority Tier",
        },
        {
            "Term/Metric": "L3M",
            "What It Means (Simple)": "Last 3 months trend window.",
            "Why It Matters": "Shows recent momentum and quick shifts.",
            "How To Use It Quickly": "Check L3M first for near-term action priority.",
            "Comprehensive Source Column": "All L3M Δ metric columns",
        },
        {
            "Term/Metric": "YoY",
            "What It Means (Simple)": "Year-over-year trend comparison window.",
            "Why It Matters": "Separates short-term noise from structural change.",
            "How To Use It Quickly": "Use with L3M to confirm sustained declines.",
            "Comprehensive Source Column": "All YoY Δ metric columns",
        },
        {
            "Term/Metric": "Partial Ratio",
            "What It Means (Simple)": "Share of mapped keywords marked as partial match.",
            "Why It Matters": "High ratio indicates stronger content-intent mismatch.",
            "How To Use It Quickly": "Reduce by expanding sections for missing intent angles.",
            "Comprehensive Source Column": "Partial Ratio",
        },
    ]
    return rows


def set_column_widths(ws) -> None:
    widths = {
        "A": 50,
        "B": 28,
        "C": 12,
        "D": 54,
        "E": 34,
        "F": 42,
        "G": 40,
        "H": 40,
        "I": 54,
        "J": 42,
        "K": 36,
        "L": 52,
        "M": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width



def main() -> None:
    with SOURCE_CSV.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        source_rows = list(reader)

    cluster_lookup = build_cluster_lookup()
    simplified = [row_to_simplified(row, cluster_lookup) for row in source_rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Simplified Opportunities"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    # Top table
    for col_idx, header in enumerate(SIMPLIFIED_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for row in simplified:
        for col_idx, header in enumerate(SIMPLIFIED_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_idx += 1

    ws.freeze_panes = "A2"
    set_column_widths(ws)

    # Glossary table start: 3 blank rows after top table
    glossary_start = row_idx + 3
    gloss_headers = [
        "Term/Metric",
        "What It Means (Simple)",
        "Why It Matters",
        "How To Use It Quickly",
        "Comprehensive Source Column",
    ]

    for col_idx, header in enumerate(gloss_headers, start=1):
        cell = ws.cell(row=glossary_start, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    g_row = glossary_start + 1
    for entry in glossary_rows():
        for col_idx, header in enumerate(gloss_headers, start=1):
            c = ws.cell(row=g_row, column=col_idx, value=entry[header])
            c.alignment = Alignment(vertical="top", wrap_text=True)
        g_row += 1

    # Slightly widen glossary text columns
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 36)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 44)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 40)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 40)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width, 48)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    # QA checks per plan
    assert len(simplified) == len(source_rows), "Row parity failed"

    src_urls = [clean(r.get("URL", "")) for r in source_rows]
    out_urls = [r["URL"] for r in simplified]
    assert src_urls == out_urls, "URL parity/order failed"

    for src, out in zip(source_rows, simplified):
        assert out["Priority"] == clean(src.get("Weighted Priority Tier", "")), "Priority parity failed"
        assert out["Primary Opportunity Keyword"] == clean(src.get("Top Keyword 1", "")), "Primary keyword parity failed"
        expected_cluster = cluster_lookup.get(normalize_url(src.get("URL", "")), clean(src.get("Topic Cluster", "")))
        assert out["Assigned Topic Cluster"] == expected_cluster, "Assigned cluster parity failed"

        expected_support = supporting_keywords(src)
        assert out["Supporting Opportunity Keywords"] == expected_support, "Supporting keywords parity failed"

        count = as_int(src.get("Metrics Available (out of 10)", ""))
        conf = clean(src.get("Data Confidence", ""))
        expected_conf = f"{conf} ({count}/10 metrics)" if conf else f"Unknown ({count}/10 metrics)"
        assert out["Data Confidence"] == expected_conf, "Confidence parity failed"

    print(f"Wrote simplified workbook: {OUTPUT_XLSX}")
    print(f"Rows: {len(simplified)} | Glossary rows: {len(glossary_rows())}")


if __name__ == "__main__":
    main()
