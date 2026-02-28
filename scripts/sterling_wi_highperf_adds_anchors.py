#!/usr/bin/env python3
"""
Sterling WI — High-performer-only ADD_CONTEXTUAL + anchor diversification.

Purpose
-------
Refines the existing WI internal-linking deliverables by:
  1) Keeping all FIX actions intact (site hygiene)
  2) Restricting ADD_CONTEXTUAL recommendations to high-performing WI content pages
     (limited to the ADD_CONTEXTUAL audited source universe):
     - GA4 Top 30 by Views
     - GSC Top 20 by Impressions
     - Ahrefs Top 20 by *live* backlinks (often 0 overlap with the audited source list; tracked as evidence)
  3) Diversifying Suggested Anchor Text via an anchor bank (15–25 per core page).

Inputs
------
  - Baseline dashboard XLSX (produced by internal_linking_dashboard_template.py)
  - Baseline implementation CSV
  - GA4 export: Pages and screens → Page path and screen class
  - GSC export: Performance → Pages (Top pages)
  - Ahrefs backlinks export (UTF-16 TSV)

Outputs (versioned; baseline preserved)
--------------------------------------
  - Updated dashboard XLSX
  - Updated implementation CSV (combined)
  - Anchor rationale CSV (explains anchor selection per ADD row)
"""

from __future__ import annotations

import argparse
import hashlib
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Any
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook


WI_BLOG_PATH_RE = re.compile(r"^/wisconsin/blog/[^/]+/?$")
WI_BLOG_URL_RE = re.compile(r"^https?://(www\.)?sterlinglawyers\.com/wisconsin/blog/[^/]+/?$", re.I)


def normalize_url(url: Any) -> str:
    value = str(url or "").strip()
    if not value or value.lower() == "nan":
        return ""

    value = value.replace("http://", "https://")
    value = value.split("#", 1)[0].strip()
    parsed = urlparse(value)

    scheme = (parsed.scheme or "https").lower()
    host = (parsed.netloc or "").lower()
    if host == "sterlinglawyers.com":
        host = "www.sterlinglawyers.com"

    path = re.sub(r"/+", "/", parsed.path or "/")
    if path and not path.endswith("/") and "." not in path.split("/")[-1]:
        path += "/"

    query = f"?{parsed.query}" if parsed.query else ""
    return f"{scheme}://{host}{path}{query}"


def is_wi_blog_url(url: Any) -> bool:
    return bool(WI_BLOG_URL_RE.match(str(url or "").strip()))


def is_wi_blog_path(path: Any) -> bool:
    value = str(path or "").strip()
    if value in ("/wisconsin/blog", "/wisconsin/blog/"):
        return False
    return bool(WI_BLOG_PATH_RE.match(value))


def stable_hash_int(text: str) -> int:
    digest = hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()
    return int(digest[:12], 16)


def read_ga4_top30(ga4_path: str, *, eligible_norm: set[str] | None = None, top_n: int = 30) -> pd.DataFrame:
    ga4 = pd.read_csv(ga4_path, comment="#")
    required = {"Page path and screen class", "Views"}
    missing = [c for c in required if c not in ga4.columns]
    if missing:
        raise ValueError(f"GA4 export missing columns: {missing}")

    ga4["GA4 Views"] = pd.to_numeric(ga4["Views"], errors="coerce").fillna(0).astype(int)
    ga4["GA4 Engaged sessions"] = pd.to_numeric(ga4.get("Engaged sessions", 0), errors="coerce").fillna(0).astype(int)

    ga4["Source URL"] = "https://www.sterlinglawyers.com" + ga4["Page path and screen class"].astype(str)
    ga4["Source URL (Normalized)"] = ga4["Source URL"].map(normalize_url)
    if eligible_norm is not None:
        ga4 = ga4[ga4["Source URL (Normalized)"].isin(eligible_norm)].copy()

    top = ga4.sort_values("GA4 Views", ascending=False).head(top_n).reset_index(drop=True)
    top["GA4 Rank (Views)"] = top.index + 1
    return top[["Source URL", "Source URL (Normalized)", "GA4 Views", "GA4 Engaged sessions", "GA4 Rank (Views)"]]


def read_gsc_top20(gsc_path: str, *, eligible_norm: set[str] | None = None, top_n: int = 20) -> pd.DataFrame:
    gsc = pd.read_csv(gsc_path)
    required = {"Top pages", "Impressions"}
    missing = [c for c in required if c not in gsc.columns]
    if missing:
        raise ValueError(f"GSC export missing columns: {missing}")

    gsc["GSC Impressions"] = pd.to_numeric(gsc["Impressions"], errors="coerce").fillna(0).astype(int)
    gsc["GSC Clicks"] = pd.to_numeric(gsc.get("Clicks", 0), errors="coerce").fillna(0).astype(int)
    gsc["GSC Position"] = pd.to_numeric(gsc.get("Position", 0), errors="coerce").fillna(0).astype(float)

    gsc["Source URL"] = gsc["Top pages"].astype(str)
    gsc["Source URL (Normalized)"] = gsc["Source URL"].map(normalize_url)
    if eligible_norm is not None:
        gsc = gsc[gsc["Source URL (Normalized)"].isin(eligible_norm)].copy()

    top = gsc.sort_values("GSC Impressions", ascending=False).head(top_n).reset_index(drop=True)
    top["GSC Rank (Impressions)"] = top.index + 1
    return top[
        ["Source URL", "Source URL (Normalized)", "GSC Impressions", "GSC Clicks", "GSC Position", "GSC Rank (Impressions)"]
    ]


def read_ahrefs_top20_live(ahrefs_path: str, *, eligible_norm: set[str] | None = None, top_n: int = 20) -> pd.DataFrame:
    ahrefs = pd.read_csv(ahrefs_path, sep="\t", encoding="utf-16")
    required = {"Referring page URL", "Target URL", "Nofollow", "Is spam", "Lost", "Redirect Chain URLs"}
    missing = [c for c in required if c not in ahrefs.columns]
    if missing:
        raise ValueError(f"Ahrefs export missing columns: {missing}")

    def canonical_target(row: pd.Series) -> str:
        target = str(row.get("Target URL", "")).strip()
        chain = str(row.get("Redirect Chain URLs", "")).strip()
        if chain and chain.lower() != "nan":
            parts = [p.strip() for p in chain.split(",") if p.strip()]
            if parts:
                return parts[-1]
        return target

    ahrefs["Canonical Target URL"] = ahrefs.apply(canonical_target, axis=1)
    ahrefs["Source URL (Normalized)"] = ahrefs["Canonical Target URL"].map(normalize_url)

    lost_nonempty = (
        ahrefs["Lost"].notna()
        & (ahrefs["Lost"].astype(str).str.strip() != "")
        & (ahrefs["Lost"].astype(str).str.lower() != "nan")
    )
    live = ahrefs[~lost_nonempty].copy()
    if eligible_norm is not None:
        live = live[live["Source URL (Normalized)"].isin(eligible_norm)].copy()
    else:
        live = live[live["Canonical Target URL"].map(is_wi_blog_url)].copy()

    def ref_domain(ref_url: Any) -> str:
        try:
            return urlparse(str(ref_url)).netloc.lower()
        except Exception:
            return ""

    live["Ref Domain"] = live["Referring page URL"].map(ref_domain)

    # Quality-weighted tie-breaker (soft signals; no hard spam exclusion):
    # - Followed (Nofollow=False): 1.0
    # - Nofollow (Nofollow=True): 0.5
    # - Spam multiplier: 0.2 if Is spam True else 1.0
    def link_weight(row: pd.Series) -> float:
        base = 1.0 if row.get("Nofollow") is False else 0.5
        spam_mult = 0.2 if row.get("Is spam") is True else 1.0
        return base * spam_mult

    live["Quality Weight"] = live.apply(link_weight, axis=1)

    agg = (
        live.groupby("Source URL (Normalized)")
        .agg(
            live_backlinks=("Source URL (Normalized)", "size"),
            referring_domains=("Ref Domain", "nunique"),
            dofollow_backlinks=("Nofollow", lambda s: int((s == False).sum())),  # noqa: E712
            nofollow_backlinks=("Nofollow", lambda s: int((s == True).sum())),  # noqa: E712
            spam_flagged_backlinks=("Is spam", lambda s: int((s == True).sum())),  # noqa: E712
            non_spam_backlinks=("Is spam", lambda s: int((s == False).sum())),  # noqa: E712
            quality_weighted_backlinks=("Quality Weight", "sum"),
        )
        .reset_index()
        .rename(
            columns={
                "live_backlinks": "Live backlinks",
                "referring_domains": "Referring domains",
                "dofollow_backlinks": "Dofollow backlinks",
                "nofollow_backlinks": "Nofollow backlinks",
                "spam_flagged_backlinks": "Spam-flagged backlinks",
                "non_spam_backlinks": "Non-spam backlinks",
                "quality_weighted_backlinks": "Quality-weighted backlinks",
            }
        )
    )

    agg = agg.sort_values(
        ["Referring domains", "Quality-weighted backlinks", "Dofollow backlinks", "Live backlinks"],
        ascending=False,
    ).head(top_n)
    agg["Ahrefs Rank (Backlinks)"] = range(1, len(agg) + 1)
    agg["Source URL"] = agg["Source URL (Normalized)"]
    return agg


@dataclass(frozen=True)
class AnchorOption:
    core_page_name: str
    category: str
    anchor: str
    rationale: str


def build_wi_anchor_bank() -> list[AnchorOption]:
    anchors: list[AnchorOption] = []

    def add(core: str, category: str, anchor: str, rationale: str) -> None:
        anchors.append(AnchorOption(core, category, anchor, rationale))

    primary = "Primary / High intent"
    secondary = "Secondary / Partial"
    variant = "Variant"
    informational = "Informational"
    additional = "Additional"

    # Divorce
    core = "Divorce"
    for a in [
        "divorce in Wisconsin",
        "Wisconsin divorce",
        "Wisconsin divorce lawyer",
        "divorce attorney in Wisconsin",
        "Wisconsin divorce attorneys",
        "filing for divorce in Wisconsin",
        "help with a divorce in Wisconsin",
    ]:
        add(core, primary, a, "High-intent service anchor (lawyer/attorney + Wisconsin where natural) to reinforce the core service page.")
    for a in [
        "the divorce process in Wisconsin",
        "how divorce works in Wisconsin",
        "getting divorced in Wisconsin",
        "Wisconsin divorce requirements",
        "grounds for divorce in Wisconsin",
        "contested divorce in Wisconsin",
        "uncontested divorce options",
        "legal steps for divorce",
        "divorce mediation in Wisconsin",
    ]:
        add(core, secondary, a, "Partial-match/process anchor to keep the in-body link natural while still signaling the core topic + Wisconsin.")
    for a in ["divorce lawyer", "divorce attorney", "divorce help", "end a marriage"]:
        add(core, variant, a, "Short variant anchor to diversify wording and avoid repetitive exact-match footprints; context carries geography.")
    for a in ["what to expect during a divorce", "divorce timelines in Wisconsin", "divorce filing procedures", "dividing property in divorce"]:
        add(core, informational, a, "Informational anchor that fits blog tone while still routing equity to the relevant core service page.")
    add(core, additional, "options for resolving a divorce", "Supportive phrasing that blends into editorial paragraphs while staying topically aligned.")

    # Child Custody
    core = "Child Custody"
    for a in [
        "child custody in Wisconsin",
        "Wisconsin child custody",
        "Wisconsin child custody lawyer",
        "child custody attorney in Wisconsin",
        "Wisconsin custody attorneys",
        "help with child custody in Wisconsin",
    ]:
        add(core, primary, a, "High-intent custody service anchor to reinforce the Wisconsin child custody core page.")
    for a in [
        "legal custody and physical placement in Wisconsin",
        "Wisconsin placement schedule",
        "physical placement rights in Wisconsin",
        "modifying a custody order in Wisconsin",
        "creating a Wisconsin parenting plan",
        "visitation and placement issues",
        "child custody and placement decisions",
        "sole custody considerations in Wisconsin",
    ]:
        add(core, secondary, a, "Wisconsin-specific partial-match anchor (placement/legal custody) to keep language accurate and natural in-body.")
    for a in ["child custody lawyer", "custody and placement", "parenting plan help", "placement factors"]:
        add(core, variant, a, "Shorter variant phrasing to diversify anchor text while remaining custody-intent.")
    for a in ["how courts decide custody and placement", "what custody arrangements can look like", "tips for a parenting plan", "what affects placement in Wisconsin"]:
        add(core, informational, a, "Informational anchor aligned to custody/placement topics; works well in narrative paragraphs.")
    add(core, additional, "support for co-parenting arrangements", "User-centric phrasing that reads naturally in editorial content and still supports the custody core page.")

    # Child Support
    core = "Child Support"
    for a in [
        "child support in Wisconsin",
        "Wisconsin child support",
        "Wisconsin child support lawyer",
        "child support attorney in Wisconsin",
        "Wisconsin child support attorneys",
        "help with child support in Wisconsin",
    ]:
        add(core, primary, a, "High-intent child support service anchor to reinforce the Wisconsin child support core page.")
    for a in [
        "Wisconsin child support guidelines",
        "calculating child support in Wisconsin",
        "modifying a child support order in Wisconsin",
        "enforcing child support in Wisconsin",
        "child support during divorce",
        "child support arrears in Wisconsin",
        "child support payments and income changes",
    ]:
        add(core, secondary, a, "Partial-match/process anchor to keep the link natural while signaling child support + Wisconsin.")
    for a in ["child support lawyer", "child support order", "support payments", "child support agreement"]:
        add(core, variant, a, "Short variant anchor to diversify phrasing; surrounding text provides Wisconsin context.")
    for a in ["how child support is calculated", "what affects child support payments", "understanding child support orders", "steps to modify child support"]:
        add(core, informational, a, "Informational anchor aligned to child support topics; appropriate for editorial paragraphs.")
    add(core, additional, "help reaching a fair child support outcome", "Outcome-focused phrasing that supports user intent without over-optimizing.")

    # Spousal Support / Alimony (Maintenance)
    core = "Alimony / Spousal Support"
    for a in [
        "spousal support in Wisconsin",
        "Wisconsin spousal support lawyer",
        "spousal maintenance in Wisconsin",
        "Wisconsin maintenance lawyer",
        "alimony attorney in Wisconsin",
    ]:
        add(core, primary, a, "High-intent service anchor using Wisconsin terminology (maintenance/spousal support) to reinforce the core page.")
    for a in [
        "how maintenance is determined in Wisconsin",
        "modifying maintenance payments in Wisconsin",
        "maintenance duration in Wisconsin",
        "temporary maintenance considerations",
        "spousal support agreements",
    ]:
        add(core, secondary, a, "Partial-match/process anchor for maintenance/spousal support that reads naturally in-body.")
    for a in ["maintenance payments", "spousal support lawyer", "alimony and maintenance", "support after divorce"]:
        add(core, variant, a, "Short variant anchor to diversify phrasing while keeping maintenance intent.")
    for a in ["what affects maintenance awards", "maintenance timelines and duration", "understanding spousal support"]:
        add(core, informational, a, "Informational anchor aligned to maintenance topics; useful for blog-style writing.")
    add(core, additional, "guidance on spousal maintenance issues", "Supportive phrasing that blends into editorial paragraphs while staying topically aligned.")

    # Property Division
    core = "Property Division"
    for a in [
        "property division in Wisconsin",
        "Wisconsin property division lawyer",
        "dividing marital property in Wisconsin",
        "splitting assets in a Wisconsin divorce",
        "Wisconsin marital property rules",
    ]:
        add(core, primary, a, "High-intent service anchor to reinforce the Wisconsin property division core page.")
    for a in [
        "marital vs non-marital property in Wisconsin",
        "dividing retirement accounts in divorce",
        "how Wisconsin courts divide property",
        "handling hidden assets in a divorce",
        "dividing a business during divorce",
        "property division agreements",
    ]:
        add(core, secondary, a, "Partial-match/process anchor that keeps property-division links natural and context-appropriate.")
    for a in ["property division lawyer", "division of assets", "marital property", "splitting property"]:
        add(core, variant, a, "Short variant anchor to diversify wording while remaining property-division intent.")
    for a in ["what counts as marital property", "tips for dividing assets fairly", "how to handle property in divorce"]:
        add(core, informational, a, "Informational anchor aligned to property division; fits editorial paragraphs.")
    add(core, additional, "help protecting property interests", "User-centric phrasing that supports conversion intent without sounding forced.")

    # Paternity
    core = "Paternity"
    for a in [
        "paternity in Wisconsin",
        "Wisconsin paternity lawyer",
        "establishing paternity in Wisconsin",
        "paternity attorney in Wisconsin",
        "help with paternity cases in Wisconsin",
    ]:
        add(core, primary, a, "High-intent paternity service anchor to reinforce the Wisconsin paternity core page.")
    for a in [
        "how to establish paternity in Wisconsin",
        "paternity and child support in Wisconsin",
        "paternity rights for fathers in Wisconsin",
        "voluntary acknowledgment of paternity",
        "paternity testing and next steps",
    ]:
        add(core, secondary, a, "Partial-match/process anchor to keep paternity links natural while signaling the core topic + Wisconsin.")
    for a in ["paternity lawyer", "paternity case", "father's rights and paternity", "paternity order"]:
        add(core, variant, a, "Short variant anchor to diversify wording; surrounding text provides context.")
    for a in ["what paternity means for custody and support", "steps to establish paternity", "questions about paternity rights"]:
        add(core, informational, a, "Informational anchor aligned to paternity; fits blog tone while routing equity to the service page.")
    add(core, additional, "support for establishing parental rights", "User-centric phrasing that aligns with paternity outcomes and reads naturally.")

    # Guardianship
    core = "Guardianship"
    for a in [
        "guardianship in Wisconsin",
        "Wisconsin guardianship lawyer",
        "guardianship attorney in Wisconsin",
        "help with guardianship in Wisconsin",
        "Wisconsin guardianship laws",
    ]:
        add(core, primary, a, "High-intent guardianship service anchor to reinforce the Wisconsin guardianship core page.")
    for a in [
        "how to become a legal guardian in Wisconsin",
        "guardianship forms and requirements",
        "temporary guardianship options",
        "guardianship for grandparents",
        "guardianship of a minor in Wisconsin",
    ]:
        add(core, secondary, a, "Partial-match/process anchor to keep guardianship links natural and helpful.")
    for a in ["legal guardian", "guardianship case", "guardian rights", "guardianship help"]:
        add(core, variant, a, "Short variant anchor to diversify wording; context supplies Wisconsin framing.")
    for a in ["what guardianship means", "steps to file for guardianship", "when guardianship is needed"]:
        add(core, informational, a, "Informational anchor aligned to guardianship; appropriate for editorial paragraphs.")
    add(core, additional, "guidance through the guardianship process", "Supportive phrasing that fits naturally in blog copy while aligning with guardianship intent.")

    return anchors


def assign_anchors(add_df: pd.DataFrame, core_url_to_name: dict[str, str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    anchor_bank = build_wi_anchor_bank()
    anchors_by_core: dict[str, list[AnchorOption]] = defaultdict(list)
    for option in anchor_bank:
        anchors_by_core[option.core_page_name].append(option)

    used_by_core: dict[str, set[str]] = defaultdict(set)
    counts_by_core_anchor: dict[str, Counter[str]] = defaultdict(Counter)

    assigned_rows: list[dict[str, Any]] = []
    rationale_rows: list[dict[str, Any]] = []

    for _, row in add_df.sort_values(["Suggested Target URL", "Source URL"]).iterrows():
        target = str(row.get("Suggested Target URL", "")).strip()
        core_name = core_url_to_name.get(normalize_url(target), "Unknown")
        bank = anchors_by_core.get(core_name, [])

        if not bank:
            anchor_text = str(row.get("Suggested Anchor Text", "")).strip()
            category = "Unknown"
            rationale = "No anchor bank available for this target; retained existing suggestion."
        else:
            start = stable_hash_int(str(row.get("Source URL", ""))) % len(bank)
            chosen: AnchorOption | None = None
            for i in range(len(bank)):
                option = bank[(start + i) % len(bank)]
                if option.anchor not in used_by_core[core_name]:
                    chosen = option
                    break
            if chosen is None:
                chosen = min(bank, key=lambda o: counts_by_core_anchor[core_name][o.anchor])

            used_by_core[core_name].add(chosen.anchor)
            counts_by_core_anchor[core_name][chosen.anchor] += 1
            anchor_text = chosen.anchor
            category = chosen.category
            rationale = chosen.rationale

        out_row = dict(row)
        out_row["Suggested Anchor Text"] = anchor_text
        assigned_rows.append(out_row)

        rationale_rows.append(
            {
                "Source URL": str(row.get("Source URL", "")).strip(),
                "Suggested Target URL": target,
                "Core Page Name": core_name,
                "Suggested Anchor Text": anchor_text,
                "Anchor Category": category,
                "Anchor Rationale": rationale,
                "Priority": str(row.get("Priority", "")).strip(),
                "Reason": str(row.get("Reason", "")).strip(),
            }
        )

    return pd.DataFrame(assigned_rows), pd.DataFrame(rationale_rows)


def write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, record in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(record, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def read_extra_urls(path: str) -> set[str]:
    df = pd.read_csv(path)
    if df.empty:
        return set()
    col = "URL" if "URL" in df.columns else df.columns[0]
    return {normalize_url(u) for u in df[col].astype(str).tolist() if normalize_url(u)}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline_xlsx", required=True)
    parser.add_argument("--baseline_impl_csv", required=True)
    parser.add_argument("--ga4_csv", required=True)
    parser.add_argument("--gsc_csv", required=True)
    parser.add_argument("--ahrefs_csv", required=True)
    parser.add_argument("--extra_hp_urls_csv", default="")
    parser.add_argument("--exclude_sources_csv", default="")
    parser.add_argument("--out_xlsx", required=True)
    parser.add_argument("--out_impl_csv", required=True)
    parser.add_argument("--out_anchor_rationale_csv", required=True)
    args = parser.parse_args()

    baseline_impl = pd.read_csv(args.baseline_impl_csv)
    if "Action" not in baseline_impl.columns:
        raise ValueError("Baseline implementation CSV missing 'Action' column.")

    add_all = baseline_impl[baseline_impl["Action"].eq("ADD_CONTEXTUAL")].copy()
    fix_all = baseline_impl[~baseline_impl["Action"].eq("ADD_CONTEXTUAL")].copy()

    audited_sources = sorted(set(add_all["Source URL"].astype(str)))
    audited_norm = {normalize_url(u) for u in audited_sources}

    ga4_top = read_ga4_top30(args.ga4_csv, eligible_norm=audited_norm, top_n=30)
    gsc_top = read_gsc_top20(args.gsc_csv, eligible_norm=audited_norm, top_n=20)
    ahrefs_top = read_ahrefs_top20_live(args.ahrefs_csv, eligible_norm=audited_norm, top_n=20)

    # High performer set for this audited source universe:
    # - union(GA4 top30, GSC top20, Ahrefs top20)
    # - plus any explicitly forced URLs (extra list)
    hp_norm = (
        set(ga4_top["Source URL (Normalized)"])
        | set(gsc_top["Source URL (Normalized)"])
        | set(ahrefs_top["Source URL (Normalized)"])
    )
    extra_norm: set[str] = set()
    if str(args.extra_hp_urls_csv or "").strip():
        extra_norm = read_extra_urls(args.extra_hp_urls_csv) & audited_norm
        hp_norm |= extra_norm

    exclude_norm: set[str] = set()
    if str(args.exclude_sources_csv or "").strip():
        exclude_norm = read_extra_urls(args.exclude_sources_csv) & audited_norm
        hp_norm -= exclude_norm

    # Filter only ADD_CONTEXTUAL actions.
    add_all["Source URL (Normalized)"] = add_all["Source URL"].map(normalize_url)
    add_hp = add_all[add_all["Source URL (Normalized)"].isin(hp_norm)].copy().drop(columns=["Source URL (Normalized)"])

    core_url_to_name = {
        normalize_url("https://www.sterlinglawyers.com/wisconsin/divorce/"): "Divorce",
        normalize_url("https://www.sterlinglawyers.com/wisconsin/child-custody/"): "Child Custody",
        normalize_url("https://www.sterlinglawyers.com/wisconsin/child-support/"): "Child Support",
        normalize_url("https://www.sterlinglawyers.com/wisconsin/spousal-support/"): "Alimony / Spousal Support",
        normalize_url("https://www.sterlinglawyers.com/wisconsin/property-division/"): "Property Division",
        normalize_url("https://www.sterlinglawyers.com/wisconsin/paternity/"): "Paternity",
        normalize_url("https://www.sterlinglawyers.com/wisconsin/guardianship/"): "Guardianship",
    }

    add_hp_assigned, rationale_df = assign_anchors(add_hp, core_url_to_name)

    # Evidence table for all audited sources (keeps dashboard run reproducible).
    evidence = pd.DataFrame({"Source URL": audited_sources})
    evidence["Source URL (Normalized)"] = evidence["Source URL"].map(normalize_url)

    # GA4 (eligible rows only)
    ga4_all = pd.read_csv(args.ga4_csv, comment="#")
    ga4_all["GA4 Views"] = pd.to_numeric(ga4_all.get("Views", 0), errors="coerce").fillna(0).astype(int)
    ga4_all["GA4 Engaged sessions"] = pd.to_numeric(ga4_all.get("Engaged sessions", 0), errors="coerce").fillna(0).astype(int)
    ga4_all["Source URL"] = "https://www.sterlinglawyers.com" + ga4_all["Page path and screen class"].astype(str)
    ga4_all["Source URL (Normalized)"] = ga4_all["Source URL"].map(normalize_url)
    ga4_all = ga4_all[ga4_all["Source URL (Normalized)"].isin(audited_norm)].copy()
    ga4_all = ga4_all.groupby("Source URL (Normalized)", as_index=False)[["GA4 Views", "GA4 Engaged sessions"]].sum()
    evidence = evidence.merge(ga4_all, on="Source URL (Normalized)", how="left")
    evidence["GA4 Views"] = evidence["GA4 Views"].fillna(0).astype(int)
    evidence["GA4 Engaged sessions"] = evidence["GA4 Engaged sessions"].fillna(0).astype(int)

    # GSC (eligible rows only)
    gsc_all = pd.read_csv(args.gsc_csv)
    gsc_all["Source URL"] = gsc_all["Top pages"].astype(str)
    gsc_all["Source URL (Normalized)"] = gsc_all["Source URL"].map(normalize_url)
    gsc_all = gsc_all[gsc_all["Source URL (Normalized)"].isin(audited_norm)].copy()
    gsc_all["GSC Impressions"] = pd.to_numeric(gsc_all.get("Impressions", 0), errors="coerce").fillna(0).astype(int)
    gsc_all["GSC Clicks"] = pd.to_numeric(gsc_all.get("Clicks", 0), errors="coerce").fillna(0).astype(int)
    gsc_all["GSC Position"] = pd.to_numeric(gsc_all.get("Position", 0), errors="coerce").fillna(0).astype(float)
    gsc_all = gsc_all.groupby("Source URL (Normalized)", as_index=False).agg(
        {"GSC Impressions": "sum", "GSC Clicks": "sum", "GSC Position": "mean"}
    )
    evidence = evidence.merge(gsc_all, on="Source URL (Normalized)", how="left")
    evidence["GSC Impressions"] = evidence["GSC Impressions"].fillna(0).astype(int)
    evidence["GSC Clicks"] = evidence["GSC Clicks"].fillna(0).astype(int)
    evidence["GSC Position"] = evidence["GSC Position"].fillna(0.0).astype(float)

    evidence["GA4 Top30 (Views)"] = evidence["Source URL (Normalized)"].isin(set(ga4_top["Source URL (Normalized)"]))
    evidence["GSC Top20 (Impressions)"] = evidence["Source URL (Normalized)"].isin(set(gsc_top["Source URL (Normalized)"]))
    evidence["Ahrefs Top20 (Live backlinks)"] = evidence["Source URL (Normalized)"].isin(
        set(ahrefs_top["Source URL (Normalized)"])
    )
    if extra_norm:
        evidence["Extra HP (Forced)"] = evidence["Source URL (Normalized)"].isin(extra_norm)
    if exclude_norm:
        evidence["Excluded (QA)"] = evidence["Source URL (Normalized)"].isin(exclude_norm)
    evidence["Is High Performer (Union)"] = evidence["Source URL (Normalized)"].isin(hp_norm)
    evidence = evidence.drop(columns=["Source URL (Normalized)"])

    # Anchor bank sheet data
    anchor_bank_df = pd.DataFrame([a.__dict__ for a in build_wi_anchor_bank()]).rename(
        columns={
            "core_page_name": "Core Page Name",
            "category": "Anchor Category",
            "anchor": "Anchor Text",
            "rationale": "Anchor Rationale",
        }
    )

    # New implementation CSV
    impl_new = pd.concat([fix_all, add_hp_assigned], ignore_index=True)

    # QA checks (hard fails)
    if len(add_hp_assigned) != add_hp_assigned["Source URL"].nunique():
        raise ValueError("ADD_CONTEXTUAL rows have duplicate Source URLs after filtering.")

    # Ensure FIX rows unchanged (row-level equality on key columns)
    baseline_fix = baseline_impl[~baseline_impl["Action"].eq("ADD_CONTEXTUAL")].copy().reset_index(drop=True)
    new_fix = impl_new[~impl_new["Action"].eq("ADD_CONTEXTUAL")].copy().reset_index(drop=True)
    if not baseline_fix.equals(new_fix):
        raise ValueError("Non-ADD actions changed; expected FIX actions to remain identical.")

    # Anchor diversity check: no single anchor >20% per target (when n>=5)
    for target, grp in add_hp_assigned.groupby("Suggested Target URL"):
        counts = grp["Suggested Anchor Text"].astype(str).value_counts()
        top_share = float(counts.iloc[0]) / float(len(grp)) if len(grp) else 0.0
        if top_share > 0.2 and len(grp) >= 5:
            raise ValueError(f"Anchor diversity failure for target {target}: top anchor share {top_share:.2%} (>20%).")

    impl_new.to_csv(args.out_impl_csv, index=False)
    rationale_df.to_csv(args.out_anchor_rationale_csv, index=False)

    # Update workbook
    wb = load_workbook(args.baseline_xlsx)

    # Update Recommendations (Add Links)
    ws_add = wb["Recommendations (Add Links)"]
    add_cols = [
        "Action",
        "Source URL",
        "Current Target URL",
        "Suggested Target URL",
        "Suggested Anchor Text",
        "Link Position",
        "Priority",
        "Reason",
    ]
    write_df_to_sheet(ws_add, add_hp_assigned[add_cols].copy())

    # Replace hidden evidence tabs
    highperf_sheet_name = "HighPerf (GA4+GSC+Ahrefs)"
    for sheet_name in [highperf_sheet_name, "Anchor Bank", "Ahrefs Top Targets (Live)"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws_hp = wb.create_sheet(highperf_sheet_name)
    write_df_to_sheet(
        ws_hp,
        evidence[
            [
                "Source URL",
                "GA4 Views",
                "GA4 Engaged sessions",
                "GSC Impressions",
                "GSC Clicks",
                "GSC Position",
                "GA4 Top30 (Views)",
                "GSC Top20 (Impressions)",
                "Ahrefs Top20 (Live backlinks)",
                *([] if "Extra HP (Forced)" not in evidence.columns else ["Extra HP (Forced)"]),
                *([] if "Excluded (QA)" not in evidence.columns else ["Excluded (QA)"]),
                "Is High Performer (Union)",
            ]
        ].copy(),
    )
    ws_hp.sheet_state = "hidden"

    ws_ab = wb.create_sheet("Anchor Bank")
    write_df_to_sheet(ws_ab, anchor_bank_df[["Core Page Name", "Anchor Category", "Anchor Text", "Anchor Rationale"]].copy())
    ws_ab.sheet_state = "hidden"

    ws_ah = wb.create_sheet("Ahrefs Top Targets (Live)")
    ah_out = ahrefs_top.copy()
    ah_out["Source URL"] = ah_out["Source URL (Normalized)"]
    ah_out["In Audited Source Set"] = ah_out["Source URL (Normalized)"].isin(audited_norm)
    ah_out = ah_out.drop(columns=["Source URL (Normalized)"])
    ordered_cols = ["Source URL"] + [c for c in ah_out.columns if c != "Source URL"]
    write_df_to_sheet(ws_ah, ah_out[ordered_cols])
    ws_ah.sheet_state = "hidden"

    # Dashboard updates
    ws_dash = wb["Dashboard"]
    ws_dash.cell(row=20, column=2, value=int(len(add_hp_assigned)))

    if ws_dash.cell(row=11, column=1).value is None and ws_dash.cell(row=11, column=2).value is None:
        ws_dash.cell(row=11, column=1, value="Add-link scope")
        ws_dash.cell(
            row=11,
            column=2,
            value="High performers only: GA4 Top30 (Views) + GSC Top20 (Impressions) + Ahrefs Top20 (Live backlinks), last 12 months",
        )

    add_counts = add_hp_assigned.groupby(add_hp_assigned["Suggested Target URL"].map(normalize_url)).size().to_dict()
    for r in range(24, 31):
        core_url = str(ws_dash.cell(row=r, column=2).value or "")
        ws_dash.cell(row=r, column=17, value=int(add_counts.get(normalize_url(core_url), 0)))

    wb.save(args.out_xlsx)


if __name__ == "__main__":
    main()
