#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urlsplit, urlunsplit

import pandas as pd


SF_MIN_COLUMNS = [
    "Type",
    "From",
    "To",
    "Anchor Text",
    "Status Code",
    "Link Position",
]


def _is_nan_like(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    s = str(value).strip().lower()
    return s in {"", "nan", "none", "null"}


def clean_url(value: Any) -> str:
    if _is_nan_like(value):
        return ""
    s = str(value).strip().strip('"').strip("'")
    s = re.sub(r"\s+", "", s)
    if s.startswith("//"):
        s = "https:" + s
    return s


def normalize_url(value: Any, *, add_trailing_slash: bool = True) -> str:
    url = clean_url(value)
    if not url:
        return ""
    if url.startswith("/"):
        # Best-effort absolute; the caller should still validate internal domains.
        url = "https://www.example.com" + url

    parts = urlsplit(url)
    scheme = (parts.scheme or "https").lower()
    netloc = parts.netloc.lower()
    path = parts.path or "/"

    if add_trailing_slash:
        last_segment = path.rsplit("/", 1)[-1]
        has_extension = "." in last_segment and not last_segment.startswith(".")
        if path != "/" and not path.endswith("/") and not has_extension:
            path = path + "/"

    return urlunsplit((scheme, netloc, path, parts.query, ""))


def url_path(value: Any) -> str:
    u = normalize_url(value, add_trailing_slash=False)
    if not u:
        return ""
    return urlsplit(u).path or "/"


def first_path_segment(value: Any) -> str:
    p = url_path(value).strip("/")
    if not p:
        return ""
    return p.split("/", 1)[0].lower()


def infer_internal_domain(core_pages: pd.DataFrame) -> str:
    if "URL" not in core_pages.columns or core_pages.empty:
        return ""
    netlocs = core_pages["URL"].map(lambda u: urlsplit(normalize_url(u)).netloc.lower()).tolist()
    netlocs = [n for n in netlocs if n]
    if not netlocs:
        return ""
    return pd.Series(netlocs).value_counts().idxmax()

def _strip_www(netloc: str) -> str:
    n = (netloc or "").strip().lower()
    return n[4:] if n.startswith("www.") else n


def canonicalize_internal_netloc(url_norm: str, *, internal_domain: str) -> str:
    """
    Ensures internal URLs use a consistent netloc (e.g., www vs non-www),
    so reporting doesn't split across domain variants.
    """
    if not url_norm or not internal_domain:
        return url_norm
    try:
        parts = urlsplit(url_norm)
    except Exception:
        return url_norm
    if not parts.netloc:
        return url_norm
    internal_netloc = urlsplit(normalize_url(internal_domain)).netloc.lower()
    if not internal_netloc:
        return url_norm
    if _strip_www(parts.netloc) != _strip_www(internal_netloc):
        return url_norm
    if parts.netloc.lower() == internal_netloc:
        return url_norm
    return urlunsplit((parts.scheme, internal_netloc, parts.path, parts.query, ""))


def is_internal(url_norm: str, *, internal_domain: str) -> bool:
    if not url_norm:
        return False
    try:
        parts = urlsplit(url_norm)
    except Exception:
        return False
    if not parts.netloc:
        return False
    if internal_domain:
        return _strip_www(parts.netloc) == _strip_www(internal_domain)
    return True


def link_position_bucket(value: Any) -> str:
    s = str(value).strip()
    return "Content" if s.lower() == "content" else "Template"


def read_csv_str(path: str | Path, *, nrows: Optional[int] = None) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False, nrows=nrows, encoding_errors="replace")


def ensure_sf_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in SF_MIN_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out


def normalize_sf_links(df: pd.DataFrame, *, internal_domain: str, redirect_map: Dict[str, str]) -> pd.DataFrame:
    out = ensure_sf_columns(df)
    out = out.copy()
    out["From"] = out["From"].map(normalize_url).map(lambda u: canonicalize_internal_netloc(u, internal_domain=internal_domain))
    out["To"] = out["To"].map(normalize_url).map(lambda u: canonicalize_internal_netloc(u, internal_domain=internal_domain))
    out["Anchor Text"] = out["Anchor Text"].astype(str).str.strip()
    out["Link Position"] = out["Link Position"].astype(str).str.strip()
    out["Type"] = out["Type"].astype(str).str.strip()
    out["Status Code"] = out["Status Code"].astype(str).str.strip()
    out["Internal"] = out["To"].map(lambda u: is_internal(u, internal_domain=internal_domain))
    out["Template Position"] = out["Link Position"].map(link_position_bucket).map(lambda b: "Template" if b != "Content" else "")

    def _canonical(to_url: str) -> str:
        u = canonicalize_internal_netloc(normalize_url(to_url), internal_domain=internal_domain)
        dest = redirect_map.get(u, u)
        return canonicalize_internal_netloc(normalize_url(dest), internal_domain=internal_domain)

    out["Canonical Target URL"] = out["To"].map(_canonical)
    return out


def filter_hyperlinks(df: pd.DataFrame) -> pd.DataFrame:
    if "Type" not in df.columns:
        return df.iloc[0:0].copy()
    return df[df["Type"].astype(str).str.lower().eq("hyperlink")].copy()


@dataclass(frozen=True)
class CorePage:
    cluster_id: str
    name: str
    url: str
    in_scope: bool


def load_core_pages_and_subservices(path: Path) -> Tuple[pd.DataFrame, List[CorePage]]:
    df = read_csv_str(path)
    core_page_name_provided = "Core Page Name" in df.columns
    # Flexible column mapping
    rename = {}
    if "URLs" in df.columns and "URL" not in df.columns:
        rename["URLs"] = "URL"
    if "CLUSTER ID" in df.columns and "Cluster ID" not in df.columns:
        rename["CLUSTER ID"] = "Cluster ID"
    if "Practice Area Type" in df.columns and "Type" not in df.columns:
        rename["Practice Area Type"] = "Type"
    if rename:
        df = df.rename(columns=rename)

    for col in ["URL", "Cluster ID"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype(str).str.strip()

    # Tier detection
    if "Tier" not in df.columns:
        if "Type" in df.columns:
            t = df["Type"].astype(str).str.strip().str.lower()
            df["Tier"] = t.map(lambda v: "Tier-1" if v == "core" else "Tier-2")
        else:
            df["Tier"] = ""
    df["Tier"] = df["Tier"].astype(str).str.strip()

    # Name detection
    if "Core Page Name" not in df.columns:
        if "Practice Area" in df.columns:
            df["Core Page Name"] = df["Practice Area"].astype(str).str.strip()
        elif "H1" in df.columns:
            df["Core Page Name"] = df["H1"].astype(str).str.strip()
        else:
            df["Core Page Name"] = ""

    # In-scope detection
    if "In Scope" not in df.columns:
        df["In Scope"] = ""
    df["In Scope"] = df["In Scope"].astype(str).str.strip()

    df["URL"] = df["URL"].map(normalize_url)
    df["Cluster ID"] = df["Cluster ID"].astype(str).str.strip()
    df["Core Page Name"] = df["Core Page Name"].astype(str).str.strip()

    # Family-law defaults (best-effort; triggered only when no explicit "Core Page Name" column exists)
    default_names = {
        "DIV": "Divorce",
        "CH-CUS": "Child Custody",
        "CH-SUP": "Child Support",
        "SPO-SUP": "Alimony / Spousal Support",
        "PT": "Paternity",
        "PROP-DIV": "Property Division",
        "GUARD": "Guardianship",
        "PRENUP": "Prenup / Postnup",
    }
    if not core_page_name_provided:
        tier1_name = df["Tier"].eq("Tier-1") & df["Cluster ID"].isin(default_names.keys())
        df.loc[tier1_name, "Core Page Name"] = df.loc[tier1_name, "Cluster ID"].map(default_names).fillna(df.loc[tier1_name, "Core Page Name"])

    blank_name = df["Core Page Name"].eq("") & df["Cluster ID"].isin(default_names.keys())
    df.loc[blank_name, "Core Page Name"] = df.loc[blank_name, "Cluster ID"].map(default_names).fillna("")

    # Core pages list (Tier-1 only)
    core_pages: List[CorePage] = []
    for _, r in df[df["Tier"].eq("Tier-1")].iterrows():
        cid = str(r.get("Cluster ID", "")).strip()
        url = normalize_url(r.get("URL", ""))
        name = str(r.get("Core Page Name", "")).strip() or cid or url
        in_scope = str(r.get("In Scope", "")).strip().lower() in {"y", "yes", "true", "1", "in scope"}
        # If no in-scope flags are set at all, default to True for Tier-1.
        core_pages.append(CorePage(cluster_id=cid, name=name, url=url, in_scope=in_scope))

    any_scope_flag = df["In Scope"].astype(str).str.strip().ne("").any()
    if not any_scope_flag:
        core_pages = [CorePage(cluster_id=p.cluster_id, name=p.name, url=p.url, in_scope=True) for p in core_pages]
        df.loc[df["Tier"].eq("Tier-1"), "In Scope"] = "Y"

    return df, core_pages


def load_content_pages(path: Path) -> pd.DataFrame:
    df = read_csv_str(path)
    if df.shape[1] == 1:
        df = df.iloc[:, [0]].copy()
        df.columns = ["Content URL"]
    # Flexible naming
    if "Blog URL" in df.columns and "Content URL" not in df.columns:
        df = df.rename(columns={"Blog URL": "Content URL"})
    if "WI Blogs" in df.columns and "Content URL" not in df.columns:
        df = df.rename(columns={"WI Blogs": "Content URL"})
    if "Content URL" not in df.columns:
        df["Content URL"] = ""
    df["Content URL"] = df["Content URL"].map(normalize_url)
    if "Cluster ID" not in df.columns:
        df["Cluster ID"] = ""
    df["Cluster ID"] = df["Cluster ID"].astype(str).str.strip()
    if "Intent Tag" not in df.columns:
        df["Intent Tag"] = ""
    df["Intent Tag"] = df["Intent Tag"].astype(str).str.strip()
    df = df[df["Content URL"].ne("")].drop_duplicates(subset=["Content URL"]).reset_index(drop=True)
    return df


def load_internal_links_3xx(path: Path) -> Tuple[pd.DataFrame, Dict[str, str]]:
    df = read_csv_str(path)
    df = ensure_sf_columns(df)

    # Detect final URL column (client files vary)
    final_col = None
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in {"final url", "final address", "redirect url", "redirect to", "redirect (core/subservice page)"}:
            final_col = c
            break
    if final_col is None:
        # Fallback: any column containing "redirect" and "page"/"url"
        for c in df.columns:
            cl = str(c).strip().lower()
            if "redirect" in cl and ("url" in cl or "address" in cl or "page" in cl):
                final_col = c
                break
    if final_col is None:
        raise ValueError("Could not detect the final URL column in the 3xx link instances file.")

    df = df.rename(columns={final_col: "Final URL"})
    df = filter_hyperlinks(df)
    df["From"] = df["From"].map(normalize_url)
    df["To"] = df["To"].map(normalize_url)
    df["Final URL"] = df["Final URL"].map(normalize_url)
    df["Status Code"] = df["Status Code"].astype(str).str.strip()
    df["Anchor Text"] = df["Anchor Text"].astype(str).str.strip()
    df["Link Position"] = df["Link Position"].astype(str).str.strip()
    df["Bucket"] = df["Link Position"].map(link_position_bucket)

    # Build redirect map: redirecting URL -> final URL
    # If duplicates exist, keep the most common final.
    map_df = df[df["To"].ne("") & df["Final URL"].ne("")].copy()
    if map_df.empty:
        return df, {}
    counts = (
        map_df.groupby(["To", "Final URL"], dropna=False)
        .size()
        .reset_index(name="Count")
        .sort_values(["To", "Count"], ascending=[True, False])
    )
    redirect_map: Dict[str, str] = {}
    for to_val, grp in counts.groupby("To", dropna=False):
        top_final = grp.iloc[0]["Final URL"]
        redirect_map[str(to_val)] = str(top_final)
    return df, redirect_map


def load_curl_map(path: Path) -> Dict[str, Tuple[int, str]]:
    """
    TSV with 3 columns (no header):
      URL <tab> FINAL_STATUS <tab> EFFECTIVE_URL
    """
    if not path.exists():
        return {}
    df = pd.read_csv(path, sep="\t", header=None, names=["URL", "Status", "Effective"], dtype=str, keep_default_na=False)
    df["URL"] = df["URL"].map(normalize_url)
    df["Effective"] = df["Effective"].map(normalize_url)
    df["Status"] = df["Status"].astype(str).str.strip()

    mapping: Dict[str, Tuple[int, str]] = {}
    for _, r in df.iterrows():
        url = str(r.get("URL", "")).strip()
        if not url:
            continue
        try:
            code = int(str(r.get("Status", "")).strip() or "0")
        except Exception:
            code = 0
        effective = str(r.get("Effective", "")).strip()
        mapping[url] = (code, effective)
        # Convenience lookups by effective URL.
        if code == 200 and effective:
            mapping.setdefault(effective, (200, effective))
    return mapping


def redirect_map_from_curl_map(curl_map: Dict[str, Tuple[int, str]], *, internal_domain: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for url, (code, effective) in curl_map.items():
        if code != 200 or not url or not effective:
            continue
        u = canonicalize_internal_netloc(normalize_url(url), internal_domain=internal_domain)
        eff = canonicalize_internal_netloc(normalize_url(effective), internal_domain=internal_domain)
        if u and eff and u != eff:
            out[u] = eff
    return out


def load_sf_links_for_sources(
    *,
    sf_edges_path: Path,
    sources: Sequence[str],
    chunksize: int = 200_000,
) -> pd.DataFrame:
    source_set = set(sources)
    if not source_set:
        return pd.DataFrame(columns=SF_MIN_COLUMNS)

    chunks: List[pd.DataFrame] = []
    for chunk in pd.read_csv(sf_edges_path, dtype=str, keep_default_na=False, chunksize=chunksize, encoding_errors="replace"):
        if "Type" not in chunk.columns or "From" not in chunk.columns:
            continue
        # Filter early by hyperlink type
        chunk = chunk[chunk["Type"].astype(str).str.lower().eq("hyperlink")].copy()
        if chunk.empty:
            continue
        chunk["From"] = chunk["From"].map(normalize_url)
        chunk = chunk[chunk["From"].isin(source_set)].copy()
        if chunk.empty:
            continue
        chunks.append(ensure_sf_columns(chunk))
    if not chunks:
        return pd.DataFrame(columns=SF_MIN_COLUMNS)
    return pd.concat(chunks, ignore_index=True)


def _top_value_and_share(series: pd.Series) -> Tuple[str, float]:
    if series.empty:
        return ("", 0.0)
    counts = series.value_counts(dropna=False)
    top = str(counts.index[0])
    share = float(counts.iloc[0]) / float(counts.sum()) if counts.sum() else 0.0
    return (top, share)


def _priority_for_source(source_url: str, *, core_urls: set[str], content_urls: set[str], bucket: str) -> str:
    if source_url in core_urls:
        return "High"
    if source_url in content_urls:
        return "Medium" if bucket == "Content" else "Low"
    return "Low"


def infer_cluster_id_from_slug(url: str) -> str:
    p = url_path(url).lower()
    if any(k in p for k in ["child-custody", "custody", "placement", "visitation", "parenting-plan"]):
        return "CH-CUS"
    if any(k in p for k in ["child-support", "support-order", "childsupport"]):
        return "CH-SUP"
    if any(k in p for k in ["spousal-support", "alimony", "maintenance"]):
        return "SPO-SUP"
    if any(k in p for k in ["prenup", "pre-nup", "postnup", "post-nup"]):
        return "PRENUP"
    if any(k in p for k in ["property-division", "marital-property", "non-marital", "nonmarital", "hidden-asset", "hidden-assets"]):
        return "PROP-DIV"
    if "paternity" in p:
        return "PT"
    if "guardianship" in p or "guardian" in p:
        return "GUARD"
    if any(k in p for k in ["divorce", "separation", "annulment", "collaborative", "mediation"]):
        return "DIV"
    return ""


def scope_prefix_to_region_name(scope_prefix: str) -> str:
    seg = scope_prefix.strip("/").split("/", 1)[0].strip()
    if not seg:
        return ""
    return seg.replace("-", " ").title()


def stable_pick(options: List[str], key: str) -> str:
    if not options:
        return ""
    # Deterministic selection to keep outputs stable across runs.
    h = 0
    for ch in key:
        h = (h * 31 + ord(ch)) % 2**32
    return options[h % len(options)]


def detect_cross_scope(url: str, *, scope_prefix: str, cross_prefixes: Sequence[str]) -> bool:
    u = normalize_url(url)
    if not u:
        return False
    p = url_path(u).lower()
    if scope_prefix and p.startswith(scope_prefix.lower()):
        return False
    for pref in cross_prefixes:
        if pref and p.startswith(pref.lower()):
            return True
    return False


def suggest_cross_scope_replacement(url: str, *, scope_prefix: str, cross_prefixes: Sequence[str]) -> str:
    u = normalize_url(url)
    if not u or not scope_prefix:
        return ""
    p = url_path(u)
    for pref in cross_prefixes:
        if pref and p.lower().startswith(pref.lower()):
            new_path = scope_prefix.rstrip("/") + p[len(pref.rstrip("/")) :]
            parts = urlsplit(u)
            return urlunsplit((parts.scheme, parts.netloc, new_path, parts.query, ""))
    return ""


def build_dashboard_tables(
    *,
    core_pages: List[CorePage],
    core_pages_df: pd.DataFrame,
    content_pages_df: pd.DataFrame,
    content_outlinks_df: pd.DataFrame,
    inlinks_by_cluster: Dict[str, pd.DataFrame],
    outlinks_by_cluster: Dict[str, pd.DataFrame],
    redirects_3xx_df: pd.DataFrame,
    redirect_map: Dict[str, str],
    curl_map: Dict[str, Tuple[int, str]],
    internal_domain: str,
    scope_prefix: str,
    cross_scope_prefixes: Sequence[str],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    content_urls = set(content_pages_df["Content URL"])
    core_urls = set([p.url for p in core_pages if p.url])
    tier2_urls = set(core_pages_df[core_pages_df["Tier"].eq("Tier-2")]["URL"].map(normalize_url))
    region_name = scope_prefix_to_region_name(scope_prefix)
    tier1_by_cluster = {
        str(r.get("Cluster ID", "")).strip(): normalize_url(r.get("URL", ""))
        for _, r in core_pages_df[core_pages_df["Tier"].eq("Tier-1")].iterrows()
        if str(r.get("Cluster ID", "")).strip()
    }

    # Normalize content outlinks once; used for add-link missing detection + QA issues.
    content_outlinks_norm = normalize_sf_links(content_outlinks_df, internal_domain=internal_domain, redirect_map=redirect_map)
    content_outlinks_norm = filter_hyperlinks(content_outlinks_norm)
    content_outlinks_norm = content_outlinks_norm[content_outlinks_norm["Internal"]].copy()

    # Core outlinks evidence (deduped across identical exports)
    core_outlinks_raw_dfs: List[pd.DataFrame] = []
    seen_out_ids: set[int] = set()
    for df_out in outlinks_by_cluster.values():
        if df_out is None or df_out.empty:
            continue
        df_id = id(df_out)
        if df_id in seen_out_ids:
            continue
        seen_out_ids.add(df_id)
        core_outlinks_raw_dfs.append(df_out)

    core_outlinks_raw = pd.concat(core_outlinks_raw_dfs, ignore_index=True) if core_outlinks_raw_dfs else pd.DataFrame(columns=SF_MIN_COLUMNS)
    core_outlinks_norm = normalize_sf_links(core_outlinks_raw, internal_domain=internal_domain, redirect_map=redirect_map)
    core_outlinks_norm = filter_hyperlinks(core_outlinks_norm)
    core_outlinks_norm = core_outlinks_norm[core_outlinks_norm["Internal"]].copy()
    cluster_by_core_url = {
        canonicalize_internal_netloc(normalize_url(p.url), internal_domain=internal_domain): p.cluster_id
        for p in core_pages
        if p.cluster_id and p.url
    }
    if not core_outlinks_norm.empty:
        core_outlinks_norm["Core Cluster ID"] = core_outlinks_norm["From"].map(lambda u: cluster_by_core_url.get(str(u), ""))

    # Fix recommendations (3xx + other issues)
    fixes: List[Dict[str, Any]] = []

    # Redirect instances (3xx): dedicated export + backfilled from content/core outlinks exports
    redirects_all_rows: List[Dict[str, Any]] = []

    redirects_norm = redirects_3xx_df.copy() if redirects_3xx_df is not None else pd.DataFrame()
    if not redirects_norm.empty:
        for col in ["From", "To", "Final URL", "Status Code", "Anchor Text", "Link Position", "Bucket"]:
            if col not in redirects_norm.columns:
                redirects_norm[col] = ""
        redirects_norm["From"] = redirects_norm["From"].map(lambda u: canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain))
        redirects_norm["To"] = redirects_norm["To"].map(lambda u: canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain))
        redirects_norm["Final URL"] = redirects_norm["Final URL"].map(lambda u: canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain))
        redirects_norm["Status Code"] = redirects_norm["Status Code"].astype(str).str.strip()
        redirects_norm["Anchor Text"] = redirects_norm["Anchor Text"].astype(str).str.strip()
        redirects_norm["Link Position"] = redirects_norm["Link Position"].astype(str).str.strip()
        redirects_norm["Bucket"] = redirects_norm["Link Position"].map(link_position_bucket)
        redirects_norm["Evidence Source"] = "3xx export"
        redirects_all_rows.extend(redirects_norm.to_dict("records"))

    def _backfill_redirects(df_norm: pd.DataFrame, *, evidence: str) -> None:
        if df_norm is None or df_norm.empty:
            return
        red = df_norm[df_norm["Status Code"].astype(str).str.startswith("3")].copy()
        if red.empty:
            return
        for _, r in red.iterrows():
            src = str(r.get("From", "")).strip()
            to = str(r.get("To", "")).strip()
            if not src or not to:
                continue
            pos = str(r.get("Link Position", "")).strip()
            anchor = str(r.get("Anchor Text", "")).strip()
            status = str(r.get("Status Code", "")).strip()
            to_norm = canonicalize_internal_netloc(normalize_url(to), internal_domain=internal_domain)
            # Prefer the merged redirect_map; fallback to curl_map if present.
            final = redirect_map.get(to_norm, "")
            if not final:
                info = curl_map.get(to_norm)
                if info and info[0] == 200 and info[1]:
                    final = info[1]
            redirects_all_rows.append(
                {
                    "Type": "Hyperlink",
                    "From": canonicalize_internal_netloc(normalize_url(src), internal_domain=internal_domain),
                    "To": to_norm,
                    "Final URL": canonicalize_internal_netloc(normalize_url(final), internal_domain=internal_domain),
                    "Status Code": status,
                    "Anchor Text": anchor,
                    "Link Position": pos,
                    "Bucket": link_position_bucket(pos),
                    "Evidence Source": evidence,
                }
            )

    _backfill_redirects(content_outlinks_norm, evidence="Content outlinks export")
    _backfill_redirects(core_outlinks_norm, evidence="Core outlinks export")

    redirects_all_df = pd.DataFrame(redirects_all_rows)
    if not redirects_all_df.empty:
        for col in ["From", "To", "Final URL", "Status Code", "Anchor Text", "Link Position", "Bucket", "Evidence Source"]:
            if col not in redirects_all_df.columns:
                redirects_all_df[col] = ""

        redirects_all_df["From"] = redirects_all_df["From"].map(lambda u: canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain))
        redirects_all_df["To"] = redirects_all_df["To"].map(lambda u: canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain))

        def _resolve_final(to_url: str, final_url: str) -> str:
            to_norm = canonicalize_internal_netloc(normalize_url(to_url), internal_domain=internal_domain)
            mapped = redirect_map.get(to_norm, "")
            if not mapped:
                info = curl_map.get(to_norm)
                if info and info[0] == 200 and info[1]:
                    mapped = info[1]
            mapped = canonicalize_internal_netloc(normalize_url(mapped), internal_domain=internal_domain)
            final_norm = canonicalize_internal_netloc(normalize_url(final_url), internal_domain=internal_domain)
            return mapped or final_norm

        redirects_all_df["Final URL"] = redirects_all_df.apply(lambda r: _resolve_final(r["To"], r["Final URL"]), axis=1)
        redirects_all_df["Bucket"] = redirects_all_df["Link Position"].map(link_position_bucket)
        redirects_all_df = redirects_all_df.drop_duplicates(
            subset=["From", "To", "Final URL", "Status Code", "Anchor Text", "Link Position"]
        ).reset_index(drop=True)
    else:
        redirects_all_df = pd.DataFrame(columns=["From", "To", "Final URL", "Status Code", "Anchor Text", "Link Position", "Bucket", "Evidence Source"])

    # Redirect fixes from redirect instances
    if not redirects_all_df.empty:
        for _, r in redirects_all_df.iterrows():
            src = normalize_url(r.get("From", ""))
            to = normalize_url(r.get("To", ""))
            final = normalize_url(r.get("Final URL", ""))
            if not src or not to:
                continue
            bucket = str(r.get("Bucket", "")).strip() or link_position_bucket(r.get("Link Position", ""))
            fixes.append(
                {
                    "Action": "FIX_REDIRECT",
                    "Source URL": src,
                    "Current Target URL": to,
                    "Suggested Target URL": final,
                    "Suggested Anchor Text": "",
                    "Link Position": str(r.get("Link Position", "")).strip(),
                    "Priority": _priority_for_source(src, core_urls=core_urls, content_urls=content_urls, bucket=bucket),
                    "Reason": f"3xx internal link ({str(r.get('Evidence Source', '')).strip() or 'redirect evidence'})",
                }
            )

    # 4xx/cross-scope/malformed from content outlinks (Content + Template)
    def _add_issue_fix(row: pd.Series, *, source_type: str) -> None:
        src = normalize_url(row.get("From", ""))
        tgt = normalize_url(row.get("To", ""))
        canon = normalize_url(row.get("Canonical Target URL", ""))
        status = str(row.get("Status Code", "")).strip()
        pos = str(row.get("Link Position", "")).strip()
        bucket = link_position_bucket(pos)
        if not src or not tgt:
            return

        # Skip redirects; handled by the 3xx report.
        if status.startswith("3"):
            return

        if status.startswith("4"):
            fixes.append(
                {
                    "Action": "REMOVE_OR_REPLACE",
                    "Source URL": src,
                    "Current Target URL": canon or tgt,
                    "Suggested Target URL": "",
                    "Suggested Anchor Text": "",
                    "Link Position": pos,
                    "Priority": _priority_for_source(src, core_urls=core_urls, content_urls=content_urls, bucket=bucket),
                    "Reason": f"{source_type}: 4xx target",
                }
            )
            return

        if detect_cross_scope(canon or tgt, scope_prefix=scope_prefix, cross_prefixes=cross_scope_prefixes):
            candidate = suggest_cross_scope_replacement(canon or tgt, scope_prefix=scope_prefix, cross_prefixes=cross_scope_prefixes)
            candidate_norm = canonicalize_internal_netloc(normalize_url(candidate), internal_domain=internal_domain)
            suggested = ""
            info = curl_map.get(candidate_norm)
            if candidate_norm and info and info[0] == 200:
                suggested = canonicalize_internal_netloc(normalize_url(info[1] or candidate_norm), internal_domain=internal_domain)
            else:
                p = url_path(canon or tgt).lower()
                if "/attorneys/" in p and internal_domain and scope_prefix:
                    suggested = normalize_url(f"https://{internal_domain}{scope_prefix.rstrip('/')}/attorneys/")
                else:
                    cid = infer_cluster_id_from_slug(canon or tgt)
                    suggested = tier1_by_cluster.get(cid, "") or next((p.url for p in core_pages if p.in_scope), "")
            fixes.append(
                {
                    "Action": "REPLACE_URL",
                    "Source URL": src,
                    "Current Target URL": canon or tgt,
                    "Suggested Target URL": suggested,
                    "Suggested Anchor Text": "",
                    "Link Position": pos,
                    "Priority": _priority_for_source(src, core_urls=core_urls, content_urls=content_urls, bucket=bucket),
                    "Reason": f"{source_type}: cross-scope link",
                }
            )
            return

        low = (canon or tgt).lower()
        if "%5bactual%20attorney%20url%5d" in low or "divorce%20lawyers%20" in low:
            fixes.append(
                {
                    "Action": "REMOVE_OR_REPLACE",
                    "Source URL": src,
                    "Current Target URL": canon or tgt,
                    "Suggested Target URL": "",
                    "Suggested Anchor Text": "",
                    "Link Position": pos,
                    "Priority": _priority_for_source(src, core_urls=core_urls, content_urls=content_urls, bucket=bucket),
                    "Reason": f"{source_type}: placeholder/malformed URL",
                }
            )

    for _, r in content_outlinks_norm.iterrows():
        _add_issue_fix(r, source_type="Content outlink")

    # 4xx/cross-scope/malformed from core outlinks evidence (Content + Template)
    for _, r in core_outlinks_norm.iterrows():
        _add_issue_fix(r, source_type="Core outlink")

    fixes_df = pd.DataFrame(fixes)
    if not fixes_df.empty:
        fixes_df = fixes_df.drop_duplicates(
            subset=[
                "Action",
                "Source URL",
                "Current Target URL",
                "Suggested Target URL",
                "Link Position",
                "Reason",
            ]
        ).reset_index(drop=True)

    # Add-link recommendations (missing-only)

    tier2_by_parent: Dict[str, List[Tuple[str, str]]] = {}
    for _, r in core_pages_df[core_pages_df["Tier"].eq("Tier-2")].iterrows():
        u = normalize_url(r.get("URL", ""))
        if not u:
            continue
        # Best-effort: infer parent by matching Tier-1 prefixes.
        parent = ""
        for core in core_pages:
            if not core.cluster_id or not core.url:
                continue
            if url_path(u).lower().startswith(url_path(core.url).lower()):
                parent = core.cluster_id
                break
        if parent:
            tier2_by_parent.setdefault(parent, []).append((u, str(r.get("Core Page Name", "")).strip()))

    def pick_target_for_content(row: pd.Series) -> Tuple[str, str]:
        cid = str(row.get("Cluster ID", "")).strip()
        if not cid:
            cid = infer_cluster_id_from_slug(row.get("Content URL", ""))
        intent = str(row.get("Intent Tag", "")).strip().lower()

        # Tier-2 selection if intent matches a Tier-2 path segment (generic heuristic).
        if cid and intent and cid in tier2_by_parent:
            for u, _name in tier2_by_parent[cid]:
                if intent and intent in url_path(u).lower():
                    return (u, "Tier-2")
        if cid and cid in tier1_by_cluster:
            return (tier1_by_cluster[cid], "Tier-1")
        # No reliable target in-scope.
        return ("", "")

    # Build a lookup of existing canonical in-body targets per content page
    content_content_links = content_outlinks_norm[content_outlinks_norm["Link Position"].astype(str).str.lower().eq("content")].copy()
    existing_targets_by_content: Dict[str, set[str]] = {}
    for src, grp in content_content_links.groupby("From", dropna=False):
        existing_targets_by_content[str(src)] = set(grp["Canonical Target URL"].astype(str).tolist())

    # Treat Fix Existing actions as satisfying “missing-only” when they will result in the desired target.
    future_targets_by_content: Dict[str, set[str]] = {k: set(v) for k, v in existing_targets_by_content.items()}
    if not fixes_df.empty:
        fx = fixes_df.copy()
        for col in ["Source URL", "Suggested Target URL", "Link Position"]:
            if col not in fx.columns:
                fx[col] = ""
        fx["Source URL"] = fx["Source URL"].map(normalize_url)
        fx["Suggested Target URL"] = fx["Suggested Target URL"].map(normalize_url)
        fx["Link Position"] = fx["Link Position"].astype(str).str.strip()
        fx = fx[fx["Source URL"].isin(content_urls) & fx["Suggested Target URL"].ne("") & fx["Link Position"].str.lower().eq("content")].copy()
        for src, grp in fx.groupby("Source URL", dropna=False):
            future_targets_by_content.setdefault(str(src), set()).update(grp["Suggested Target URL"].astype(str).tolist())

    adds: List[Dict[str, Any]] = []
    for _, r in content_pages_df.iterrows():
        src = normalize_url(r.get("Content URL", ""))
        if not src:
            continue
        target, tier = pick_target_for_content(r)
        target = normalize_url(target)
        if not target:
            continue
        existing = future_targets_by_content.get(src, set())
        if target in existing:
            continue
        # Anchor: varied, human-readable, includes core term + region when natural.
        cid = str(r.get("Cluster ID", "")).strip() or infer_cluster_id_from_slug(src)
        core_name = next((p.name for p in core_pages if p.cluster_id == cid and p.name), "") or cid or "Service"
        term = core_name
        if region_name:
            term = re.sub(rf"(?i)\\b{re.escape(region_name)}\\b", "", term).strip()
        term = re.sub(r"(?i)\\blaws\\b", "", term).strip()
        term = re.sub(r"\\s+", " ", term).strip(" -–—|/")
        if not term:
            term = core_name

        if region_name:
            anchor_variants = [
                f"{region_name} {term}",
                f"{term} in {region_name}",
                f"{region_name} {term} services",
            ]
        else:
            anchor_variants = [term, f"{term} services"]
        anchor = stable_pick(anchor_variants, src)
        adds.append(
            {
                "Action": "ADD_CONTEXTUAL",
                "Source URL": src,
                "Current Target URL": "",
                "Suggested Target URL": target,
                "Suggested Anchor Text": anchor,
                "Link Position": "Content",
                "Priority": "High",
                "Reason": "Add 1 intentional in-body link (missing-only policy)",
            }
        )

    adds_df = pd.DataFrame(adds)

    # Core Pages table metrics
    rows: List[Dict[str, Any]] = []
    inbound_examples: List[Dict[str, Any]] = []
    outbound_examples: List[Dict[str, Any]] = []

    for p in core_pages:
        if not p.in_scope:
            continue
        cid = p.cluster_id
        core_url = normalize_url(p.url)

        # Inlinks
        in_df_raw = inlinks_by_cluster.get(cid, pd.DataFrame())
        in_df_norm = normalize_sf_links(in_df_raw, internal_domain=internal_domain, redirect_map=redirect_map)
        in_df_norm = filter_hyperlinks(in_df_norm)
        in_df_norm = in_df_norm[in_df_norm["Internal"]].copy()
        in_df_norm = in_df_norm[in_df_norm["Canonical Target URL"].eq(core_url)].copy() if core_url else in_df_norm

        in_instances = int(len(in_df_norm))
        unique_linking_pages = int(in_df_norm["From"].nunique()) if not in_df_norm.empty else 0
        in_content = int((in_df_norm["Link Position"].astype(str).str.lower() == "content").sum()) if not in_df_norm.empty else 0
        in_template = in_instances - in_content

        # Blog/content support
        blog_content = 0
        if not in_df_norm.empty:
            from_is_content = in_df_norm["From"].isin(content_urls)
            blog_content = int((from_is_content & (in_df_norm["Link Position"].astype(str).str.lower() == "content")).sum())

        # Anchor (Content only)
        anchor_top = ""
        anchor_share = 0.0
        if in_content > 0:
            anchors = in_df_norm[in_df_norm["Link Position"].astype(str).str.lower().eq("content")]["Anchor Text"].astype(str).str.strip()
            anchors = anchors.replace("", "(empty)")
            anchor_top, anchor_share = _top_value_and_share(anchors)

        # Inbound examples: top 3 linking pages (Content)
        if in_content > 0:
            g = in_df_norm[in_df_norm["Link Position"].astype(str).str.lower().eq("content")].copy()
            src_counts = g.groupby("From").size().reset_index(name="Count").sort_values("Count", ascending=False).head(3)
            for _, rr in src_counts.iterrows():
                src_url = rr["From"]
                sample_anchor = (
                    g[g["From"].eq(src_url)]["Anchor Text"].astype(str).replace("", "(empty)").value_counts().head(1).index.tolist()
                )
                inbound_examples.append(
                    {
                        "Core Page": p.name,
                        "Core Page URL": core_url,
                        "Linking Page": src_url,
                        "Example Anchor Text": sample_anchor[0] if sample_anchor else "",
                        "Count": int(rr["Count"]),
                        "Link Position": "Content",
                    }
                )

        # Redirecting inlinks (Content) from 3xx instances
        redirect_inlinks_content = 0
        if not redirects_all_df.empty and core_url:
            redirect_inlinks_content = int((redirects_all_df["Final URL"].eq(core_url) & redirects_all_df["Bucket"].eq("Content")).sum())

        # Outlinks
        out_df_raw = outlinks_by_cluster.get(cid, pd.DataFrame())
        out_df_norm = normalize_sf_links(out_df_raw, internal_domain=internal_domain, redirect_map=redirect_map)
        out_df_norm = filter_hyperlinks(out_df_norm)
        out_df_norm = out_df_norm[out_df_norm["Internal"]].copy()
        # Combined-export mode support: if outlinks contain multiple sources, scope to this core page.
        if core_url and out_df_norm["From"].nunique() > 1:
            out_df_norm = out_df_norm[out_df_norm["From"].eq(core_url)].copy()

        out_content_df = out_df_norm[out_df_norm["Link Position"].astype(str).str.lower().eq("content")].copy()
        out_to_tier2_content = int(out_content_df["Canonical Target URL"].isin(tier2_urls).sum()) if not out_content_df.empty else 0
        out_supporting_content = int(len(out_content_df)) - out_to_tier2_content

        # Outbound examples: top 3 canonical targets (Content)
        if not out_content_df.empty:
            tgt_counts = (
                out_content_df.groupby("Canonical Target URL")
                .size()
                .reset_index(name="Count")
                .sort_values("Count", ascending=False)
                .head(3)
            )
            for _, rr in tgt_counts.iterrows():
                tgt_url = rr["Canonical Target URL"]
                sample_anchor = (
                    out_content_df[out_content_df["Canonical Target URL"].eq(tgt_url)]["Anchor Text"]
                    .astype(str)
                    .replace("", "(empty)")
                    .value_counts()
                    .head(1)
                    .index.tolist()
                )
                outbound_examples.append(
                    {
                        "Core Page": p.name,
                        "Core Page URL": core_url,
                        "Linked-To Page": tgt_url,
                        "Example Anchor Text": sample_anchor[0] if sample_anchor else "",
                        "Count": int(rr["Count"]),
                        "Link Position": "Content",
                    }
                )

        # Issue counts (Content + Template)
        redirects_outlinks = 0
        if not redirects_all_df.empty and core_url:
            redirects_outlinks = int((redirects_all_df["From"].eq(core_url)).sum())
        out_4xx = int(out_df_norm["Status Code"].astype(str).str.startswith("4").sum()) if not out_df_norm.empty else 0
        cross_scope = int(out_df_norm["Canonical Target URL"].map(lambda u: detect_cross_scope(u, scope_prefix=scope_prefix, cross_prefixes=cross_scope_prefixes)).sum()) if not out_df_norm.empty else 0

        # Action counts per core page
        fix_count = int((fixes_df["Source URL"].eq(core_url)).sum()) if not fixes_df.empty else 0
        add_count = int((adds_df["Suggested Target URL"].eq(core_url)).sum()) if not adds_df.empty else 0

        rows.append(
            {
                "Core Page Name": p.name,
                "Canonical Core Page URL": core_url,
                "Inlink Instances": in_instances,
                "Unique Linking Pages": unique_linking_pages,
                "Inlinks (Link Position=Content)": in_content,
                "Inlinks (Template Position)": in_template,
                "Blog Inlinks (Link Position=Content)": blog_content,
                "Redirecting Inlinks (Content)": redirect_inlinks_content,
                "Top Anchor (Content)": anchor_top,
                "Top Anchor Share % (Content)": round(anchor_share * 100.0, 1) if anchor_share else 0.0,
                "Outbound Content Links → Tier-2": out_to_tier2_content,
                "Outbound Content Links → Supporting/Other": out_supporting_content,
                "Redirecting Outlinks (3xx) (All Positions)": redirects_outlinks,
                "4xx Outlinks (All Positions)": out_4xx,
                "Cross-Scope Outlinks (All Positions)": cross_scope,
                "Fix Actions (Count)": fix_count,
                "Add Actions (Count)": add_count,
            }
        )

    core_pages_table = pd.DataFrame(rows)

    # Redirect hygiene spotlight (top redirecting targets)
    redirect_spotlight = pd.DataFrame()
    if not redirects_all_df.empty:
        redirect_spotlight = (
            redirects_all_df.groupby(["To", "Final URL"], dropna=False)
            .size()
            .reset_index(name="Instances")
            .sort_values("Instances", ascending=False)
        )

    # Scope & Rules (Dashboard)
    scope_rows = [
        ("Type counted", "Hyperlink"),
        ("Buckets", "Link Position=Content vs Template Position (non-Content)"),
        ("URL used for reporting", "Canonical Normalized URL (redirect-resolved)"),
        ("Data gaps", "Prenup/Postnup excluded (not provided as audited targets in this run)."),
        ("Internal domain", internal_domain),
        ("In-scope prefix", scope_prefix or "(not set)"),
        ("Cross-scope prefixes", ", ".join(cross_scope_prefixes) if cross_scope_prefixes else "(not set)"),
    ]
    scope_df = pd.DataFrame(scope_rows, columns=["Setting", "Value"])

    # Executive KPIs
    kpis = []
    kpis.append(("Core Pages in scope", int(sum(1 for p in core_pages if p.in_scope))))
    kpis.append(("Content pages analyzed", int(len(content_pages_df))))
    kpis.append(("Redirecting link instances (3xx)", int(len(redirects_all_df))))
    if not fixes_df.empty:
        for a, c in fixes_df["Action"].value_counts().items():
            kpis.append((f"Fix actions: {a}", int(c)))
    kpis.append(("Add-link recommendations (missing-only)", int(len(adds_df))))
    kpi_df = pd.DataFrame(kpis, columns=["Metric", "Value"])

    # QA issues (evidence for fixes)
    qa_rows: List[Dict[str, Any]] = []
    if not redirects_all_df.empty:
        for _, r in redirects_all_df.iterrows():
            qa_rows.append(
                {
                    "Issue Type": "INTERNAL_REDIRECT_TARGET",
                    "Source URL": normalize_url(r.get("From", "")),
                    "Current Target URL": normalize_url(r.get("To", "")),
                    "Suggested Target URL": normalize_url(r.get("Final URL", "")),
                    "Status Code": str(r.get("Status Code", "")).strip(),
                    "Link Position": str(r.get("Link Position", "")).strip(),
                    "Anchor Text": str(r.get("Anchor Text", "")).strip(),
                    "Notes": f"Update internal link to final destination ({str(r.get('Evidence Source', '')).strip() or 'redirect evidence'}).",
                }
            )
    # Add 4xx/cross-scope evidence from content outlinks + core outlinks
    def _qa_from_df(df_norm: pd.DataFrame, *, source_label: str) -> None:
        if df_norm is None or df_norm.empty:
            return
        for _, r in df_norm.iterrows():
            status = str(r.get("Status Code", "")).strip()
            if not status or status.startswith("2") or status.startswith("3"):
                continue
            src = str(r.get("From", "")).strip()
            tgt = str(r.get("Canonical Target URL", "")).strip() or str(r.get("To", "")).strip()
            if status.startswith("4"):
                qa_rows.append(
                    {
                        "Issue Type": "INTERNAL_4XX_TARGET",
                        "Source URL": src,
                        "Current Target URL": tgt,
                        "Suggested Target URL": "",
                        "Status Code": status,
                        "Link Position": str(r.get("Link Position", "")).strip(),
                        "Anchor Text": str(r.get("Anchor Text", "")).strip(),
                        "Notes": f"{source_label}: 4xx internal link; remove or replace.",
                    }
                )
    _qa_from_df(content_outlinks_norm, source_label="Content outlinks")
    _qa_from_df(core_outlinks_norm, source_label="Core page outlinks")

    # Cross-scope evidence (internal)
    for df_norm, source_label in [(content_outlinks_norm, "Content outlinks"), (core_outlinks_norm, "Core page outlinks")]:
        if df_norm is None or df_norm.empty:
            continue
        cross = df_norm[df_norm["Canonical Target URL"].map(lambda u: detect_cross_scope(u, scope_prefix=scope_prefix, cross_prefixes=cross_scope_prefixes))]
        for _, r in cross.iterrows():
            current = str(r.get("Canonical Target URL", "")).strip()
            candidate = suggest_cross_scope_replacement(current, scope_prefix=scope_prefix, cross_prefixes=cross_scope_prefixes)
            candidate_norm = canonicalize_internal_netloc(normalize_url(candidate), internal_domain=internal_domain)
            suggested = ""
            info = curl_map.get(candidate_norm)
            if candidate_norm and info and info[0] == 200:
                suggested = canonicalize_internal_netloc(normalize_url(info[1] or candidate_norm), internal_domain=internal_domain)
            else:
                path_lower = url_path(current).lower()
                if "/attorneys/" in path_lower and internal_domain and scope_prefix:
                    suggested = normalize_url(f"https://{internal_domain}{scope_prefix.rstrip('/')}/attorneys/")
                else:
                    cid = infer_cluster_id_from_slug(current)
                    suggested = tier1_by_cluster.get(cid, "") or next((p.url for p in core_pages if p.in_scope), "")
            qa_rows.append(
                {
                    "Issue Type": "CROSS_SCOPE_LINK",
                    "Source URL": str(r.get("From", "")).strip(),
                    "Current Target URL": current,
                    "Suggested Target URL": suggested,
                    "Status Code": str(r.get("Status Code", "")).strip(),
                    "Link Position": str(r.get("Link Position", "")).strip(),
                    "Anchor Text": str(r.get("Anchor Text", "")).strip(),
                    "Notes": f"{source_label}: cross-scope internal link; replace with in-scope equivalent when possible.",
                }
            )

    qa_df = pd.DataFrame(qa_rows).drop_duplicates().reset_index(drop=True)

    # Recommendations tabs (appendices)
    rec_fix_df = fixes_df.copy()
    rec_add_df = adds_df.copy()

    # Redirects (3xx) appendix (normalized + summarized)
    redirects_sheet = redirects_all_df.copy()
    if not redirects_sheet.empty:
        redirects_sheet = redirects_sheet.rename(columns={"Final URL": "Suggested Target URL", "To": "Current Target URL"})

        def _curl_status(u: Any) -> str:
            u_norm = canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain)
            code = curl_map.get(u_norm, (0, ""))[0]
            return str(code) if code else ""

        def _curl_effective(u: Any) -> str:
            u_norm = canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain)
            eff = curl_map.get(u_norm, (0, ""))[1]
            return canonicalize_internal_netloc(normalize_url(eff), internal_domain=internal_domain) if eff else ""

        redirects_sheet["Curl Final Status"] = redirects_sheet["Current Target URL"].map(_curl_status)
        redirects_sheet["Curl Effective URL"] = redirects_sheet["Current Target URL"].map(_curl_effective)

        redirects_sheet = redirects_sheet[
            [
                "From",
                "Current Target URL",
                "Suggested Target URL",
                "Curl Final Status",
                "Curl Effective URL",
                "Status Code",
                "Anchor Text",
                "Link Position",
                "Bucket",
                "Evidence Source",
            ]
        ].rename(columns={"From": "Source URL"})

    return (
        scope_df,
        kpi_df,
        core_pages_table,
        pd.DataFrame(inbound_examples),
        pd.DataFrame(outbound_examples),
        redirect_spotlight,
        qa_df,
    ), rec_fix_df, rec_add_df, redirects_sheet, content_outlinks_norm, core_outlinks_norm


def write_workbook(
    *,
    out_xlsx: Path,
    dashboard_tables: Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame],
    rec_fix_df: pd.DataFrame,
    rec_add_df: pd.DataFrame,
    redirects_sheet: pd.DataFrame,
    content_outlinks_norm: pd.DataFrame,
    core_outlinks_norm: pd.DataFrame,
    hide_appendices: bool = True,
) -> None:
    scope_df, kpi_df, core_pages_table, inbound_examples, outbound_examples, redirect_spotlight, qa_df = dashboard_tables

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # Placeholder; we will format Dashboard with openpyxl after write.
        pd.DataFrame({"Dashboard": ["See formatted Dashboard tab."]}).to_excel(writer, index=False, sheet_name="Dashboard")
        core_pages_table.fillna("").to_excel(writer, index=False, sheet_name="Core Pages Inlinks Audit")
        core_outlinks_norm.fillna("").to_excel(writer, index=False, sheet_name="Core Pages Outlinks Audit")
        content_outlinks_norm.fillna("").to_excel(writer, index=False, sheet_name="Content Outlinks (Current)")
        redirects_sheet.fillna("").to_excel(writer, index=False, sheet_name="Redirecting Links (3xx)")
        qa_df.fillna("").to_excel(writer, index=False, sheet_name="QA Issues")
        rec_fix_df.fillna("").to_excel(writer, index=False, sheet_name="Recommendations (Fix Existing)")
        rec_add_df.fillna("").to_excel(writer, index=False, sheet_name="Recommendations (Add Links)")

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(out_xlsx)
    ws = wb["Dashboard"]
    ws.delete_rows(1, ws.max_row)

    title_font = Font(bold=True, size=16)
    section_font = Font(bold=True, size=12, color="FFFFFF")
    header_font = Font(bold=True)
    fill_section = PatternFill("solid", fgColor="2F5597")
    fill_header = PatternFill("solid", fgColor="D9E1F2")

    def write_section_title(row: int, col: int, text: str, width_cols: int = 6) -> int:
        ws.cell(row=row, column=col, value=text).font = section_font
        ws.cell(row=row, column=col).fill = fill_section
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width_cols - 1)
        return row + 1

    def write_df(start_row: int, start_col: int, df: pd.DataFrame, *, header_fill: bool = True) -> int:
        # headers
        for j, col_name in enumerate(df.columns, start=start_col):
            c = ws.cell(row=start_row, column=j, value=str(col_name))
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if header_fill:
                c.fill = fill_header
        # rows
        for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
            for j, col_name in enumerate(df.columns, start=start_col):
                val = r[col_name]
                ws.cell(row=i, column=j, value=val)
        return start_row + len(df) + 2

    # Title
    ws.cell(row=1, column=1, value="Internal Linking Audit + Topical Authority Dashboard").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    r = 3
    r = write_section_title(r, 1, "Scope & Rules", width_cols=6)
    r = write_df(r, 1, scope_df, header_fill=True)

    r = write_section_title(r, 1, "Executive KPIs", width_cols=6)
    r = write_df(r, 1, kpi_df)

    r = write_section_title(r, 1, "Core Pages (Audit Summary)", width_cols=10)
    r = write_df(r, 1, core_pages_table)
    core_table_end = r

    # Evidence tables side-by-side (Top 3 examples)
    if not inbound_examples.empty:
        r = write_section_title(core_table_end, 1, "Top Inbound Examples (Link Position=Content)", width_cols=6)
        r_in_end = write_df(r, 1, inbound_examples)
    else:
        r_in_end = core_table_end

    if not outbound_examples.empty:
        r2 = write_section_title(core_table_end, 8, "Top Outbound Examples (Link Position=Content)", width_cols=6)
        _ = write_df(r2, 8, outbound_examples)

    # Redirect spotlight
    if redirect_spotlight is not None and not redirect_spotlight.empty:
        start = max(r_in_end, core_table_end) + 1
        start = write_section_title(start, 1, "Redirect Hygiene Spotlight (Top Redirecting Targets)", width_cols=10)
        _ = write_df(start, 1, redirect_spotlight.head(15))

    # Light formatting: freeze panes and column widths
    ws.freeze_panes = "A3"
    for ws2 in wb.worksheets:
        ws2.auto_filter.ref = ws2.dimensions
        max_row = min(ws2.max_row, 2000)
        for col_idx in range(1, ws2.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, max_row + 1):
                v = ws2.cell(row=row_idx, column=col_idx).value
                s = "" if v is None else str(v)
                max_len = max(max_len, len(s))
            ws2.column_dimensions[col_letter].width = min(max(12, max_len + 2), 70)

    # Hide appendices
    if hide_appendices:
        for name in wb.sheetnames:
            if name != "Dashboard":
                wb[name].sheet_state = "hidden"

    wb.save(out_xlsx)


def write_implementation_csv(out_csv: Path, rec_fix_df: pd.DataFrame, rec_add_df: pd.DataFrame) -> pd.DataFrame:
    combined = pd.concat([rec_fix_df, rec_add_df], ignore_index=True, sort=False)
    if combined.empty:
        out_csv.parent.mkdir(parents=True, exist_ok=True)
        combined.to_csv(out_csv, index=False)
        return combined

    # Ensure required columns exist and order them
    required = [
        "Action",
        "Source URL",
        "Current Target URL",
        "Suggested Target URL",
        "Suggested Anchor Text",
        "Link Position",
        "Priority",
        "Reason",
    ]
    for c in required:
        if c not in combined.columns:
            combined[c] = ""
    combined = combined[required].copy()
    combined = combined.fillna("")
    combined = combined.drop_duplicates().reset_index(drop=True)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(out_csv, index=False)
    return combined


def main() -> int:
    parser = argparse.ArgumentParser(description="One-dashboard internal linking audit + topical authority template (segmented exports + 3xx link instances).")
    parser.add_argument("--core_pages", required=True, help="Core pages + subservices CSV (Tier-1/Tier-2, Cluster ID, URL, etc.)")
    parser.add_argument("--content_pages", required=True, help="Content/blog list CSV (Content URL, optional Cluster ID/Intent Tag)")
    parser.add_argument("--content_outlinks", required=True, help="Screaming Frog edges export for content/blog sources (or sitewide outlinks export)")
    parser.add_argument("--redirect_3xx", required=True, help="Internal redirecting link instances (3xx) CSV with From/To and Final URL")
    parser.add_argument("--core_exports_map", required=True, help="CSV mapping Cluster ID -> inlinks/outlinks file paths for Tier-1 core pages")
    parser.add_argument("--curl_map", default="", help="Optional TSV (no header): URL<TAB>FINAL_STATUS<TAB>EFFECTIVE_URL. Used for redirect resolution + QA.")
    parser.add_argument("--scope_prefix", default="", help="In-scope path prefix like '/illinois/' or '/wisconsin/' (optional, improves cross-scope checks)")
    parser.add_argument("--cross_scope_prefixes", default="/illinois/,/wisconsin/,/iowa/", help="Comma-separated cross-scope prefixes (default covers Sterling multi-state)")
    parser.add_argument("--internal_domain", default="", help="Internal domain (defaults to most common netloc in core_pages)")
    parser.add_argument("--out_xlsx", required=True)
    parser.add_argument("--out_csv", required=True)
    parser.add_argument("--hide_appendices", action="store_true", default=True)
    args = parser.parse_args()

    core_pages_df, core_pages = load_core_pages_and_subservices(Path(args.core_pages))
    content_pages_df = load_content_pages(Path(args.content_pages))
    redirects_3xx_df, redirect_map = load_internal_links_3xx(Path(args.redirect_3xx))

    internal_domain = args.internal_domain.strip() or infer_internal_domain(core_pages_df)
    if not internal_domain:
        raise ValueError("Could not infer internal domain; pass --internal_domain.")

    # Optional curl map (network should be run outside this script in restricted environments)
    curl_map: Dict[str, Tuple[int, str]] = {}
    if args.curl_map:
        raw = load_curl_map(Path(args.curl_map))
        # Canonicalize keys/values to the internal domain netloc.
        for u, (code, eff) in raw.items():
            u_norm = canonicalize_internal_netloc(normalize_url(u), internal_domain=internal_domain)
            eff_norm = canonicalize_internal_netloc(normalize_url(eff), internal_domain=internal_domain) if eff else ""
            if not u_norm:
                continue
            curl_map[u_norm] = (code, eff_norm)
            if code == 200 and eff_norm:
                curl_map.setdefault(eff_norm, (200, eff_norm))
        redirect_map.update(redirect_map_from_curl_map(curl_map, internal_domain=internal_domain))

    # Canonicalize redirect map keys/values to the internal domain netloc so lookups are stable.
    if redirect_map:
        canon: Dict[str, str] = {}
        for k, v in redirect_map.items():
            kk = canonicalize_internal_netloc(normalize_url(k), internal_domain=internal_domain)
            vv = canonicalize_internal_netloc(normalize_url(v), internal_domain=internal_domain)
            if kk and vv:
                canon[kk] = vv
        redirect_map = canon

    # Load core exports map
    map_df = read_csv_str(args.core_exports_map)
    # Flexible columns
    col_map = {}
    for c in map_df.columns:
        cl = c.strip().lower()
        if cl in {"cluster id", "cluster_id"}:
            col_map[c] = "Cluster ID"
        if cl in {"inlinks", "inlinks path", "inlinks_file", "inlinks file"}:
            col_map[c] = "Inlinks Path"
        if cl in {"outlinks", "outlinks path", "outlinks_file", "outlinks file"}:
            col_map[c] = "Outlinks Path"
    if col_map:
        map_df = map_df.rename(columns=col_map)
    for required in ["Cluster ID", "Inlinks Path", "Outlinks Path"]:
        if required not in map_df.columns:
            raise ValueError(f"core_exports_map is missing required column: {required}")

    inlinks_by_cluster: Dict[str, pd.DataFrame] = {}
    outlinks_by_cluster: Dict[str, pd.DataFrame] = {}
    csv_cache: Dict[str, pd.DataFrame] = {}
    for _, r in map_df.iterrows():
        cid = str(r.get("Cluster ID", "")).strip()
        if not cid:
            continue
        in_path = Path(str(r.get("Inlinks Path", "")).strip())
        out_path = Path(str(r.get("Outlinks Path", "")).strip())
        if in_path.exists():
            key = str(in_path)
            if key not in csv_cache:
                csv_cache[key] = read_csv_str(in_path)
            inlinks_by_cluster[cid] = csv_cache[key]
        else:
            inlinks_by_cluster[cid] = pd.DataFrame(columns=SF_MIN_COLUMNS)
        if out_path.exists():
            key = str(out_path)
            if key not in csv_cache:
                csv_cache[key] = read_csv_str(out_path)
            outlinks_by_cluster[cid] = csv_cache[key]
        else:
            outlinks_by_cluster[cid] = pd.DataFrame(columns=SF_MIN_COLUMNS)

    # Filter content outlinks to just the provided content pages (chunked)
    content_urls = content_pages_df["Content URL"].tolist()
    content_outlinks_raw = load_sf_links_for_sources(sf_edges_path=Path(args.content_outlinks), sources=content_urls)

    scope_prefix = args.scope_prefix.strip()
    if scope_prefix and not scope_prefix.startswith("/"):
        scope_prefix = "/" + scope_prefix
    if scope_prefix and not scope_prefix.endswith("/"):
        scope_prefix = scope_prefix + "/"
    cross_prefixes = [p.strip() for p in str(args.cross_scope_prefixes).split(",") if p.strip()]
    cross_prefixes = [p if p.startswith("/") else "/" + p for p in cross_prefixes]
    cross_prefixes = [p if p.endswith("/") else p + "/" for p in cross_prefixes]

    dashboard_tables, rec_fix_df, rec_add_df, redirects_sheet, content_outlinks_norm, core_outlinks_norm = build_dashboard_tables(
        core_pages=core_pages,
        core_pages_df=core_pages_df,
        content_pages_df=content_pages_df,
        content_outlinks_df=content_outlinks_raw,
        inlinks_by_cluster=inlinks_by_cluster,
        outlinks_by_cluster=outlinks_by_cluster,
        redirects_3xx_df=redirects_3xx_df,
        redirect_map=redirect_map,
        curl_map=curl_map,
        internal_domain=internal_domain,
        scope_prefix=scope_prefix,
        cross_scope_prefixes=cross_prefixes,
    )

    out_xlsx = Path(args.out_xlsx)
    out_csv = Path(args.out_csv)
    write_workbook(
        out_xlsx=out_xlsx,
        dashboard_tables=dashboard_tables,
        rec_fix_df=rec_fix_df,
        rec_add_df=rec_add_df,
        redirects_sheet=redirects_sheet,
        content_outlinks_norm=content_outlinks_norm,
        core_outlinks_norm=core_outlinks_norm,
        hide_appendices=args.hide_appendices,
    )
    write_implementation_csv(out_csv, rec_fix_df, rec_add_df)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
