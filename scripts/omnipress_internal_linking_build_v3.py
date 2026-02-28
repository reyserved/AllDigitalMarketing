#!/usr/bin/env python3
"""
Omnipress internal linking worksheet + Sterling-style Anchor Bank (v3).

Goal (v3): increase anchor text variation and reduce repetitive lead-ins like
\"how\" and \"see how\", while preserving:
  - linking rules (blog/conversion: 1 pillar + 2 cross-links; assets: 1 link)
  - per-target anchor mix enforcement for core/conversion targets
  - banned terms (best/compare) and services/solutions exceptions

Inputs (authoritative, read-only):
  - /Users/reymartjansarigumba/Desktop/.../omnipress-page-master.csv
  - /Users/reymartjansarigumba/Desktop/.../omnipress-core-pages-cluster.csv

Outputs (written under output/spreadsheet/):
  - Omnipress_Internal_Linking_Rules_Anchors_2026-02-28_v3.xlsx
  - Omnipress_ADD_CONTEXTUAL_UPDATED_2026-02-28_v3.csv
  - Omnipress_Anchor_Bank_AllTargets_2026-02-28_v3.csv
  - Omnipress_Anchor_Bank_CoreTargets_2026-02-28_v3.csv
  - Omnipress_Anchor_Bank_Assignments_2026-02-28_v3.xlsx
"""

from __future__ import annotations

import hashlib
import math
import re
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


PAGE_MASTER_PATH = (
    "/Users/reymartjansarigumba/Desktop/Antigravity/ROCKET CLICKS/ALL DIGITAL MARKETING/omnipress-page-master.csv"
)
CORE_CLUSTERS_PATH = (
    "/Users/reymartjansarigumba/Desktop/Antigravity/ROCKET CLICKS/ALL DIGITAL MARKETING/omnipress-core-pages-cluster.csv"
)

OUT_WORKBOOK = "output/spreadsheet/Omnipress_Internal_Linking_Rules_Anchors_2026-02-28_v3.xlsx"
OUT_WORKSHEET_CSV = "output/spreadsheet/Omnipress_ADD_CONTEXTUAL_UPDATED_2026-02-28_v3.csv"
OUT_ANCHOR_BANK_ALL = "output/spreadsheet/Omnipress_Anchor_Bank_AllTargets_2026-02-28_v3.csv"
OUT_ANCHOR_BANK_CORE = "output/spreadsheet/Omnipress_Anchor_Bank_CoreTargets_2026-02-28_v3.csv"
OUT_ASSIGNMENTS_XLSX = "output/spreadsheet/Omnipress_Anchor_Bank_Assignments_2026-02-28_v3.xlsx"


CORE_TARGET_TYPES = {
    "Product Hub",
    "Product",
    "Service",
    "Main Service",
    "Product Supporting",
}

ASSET_SOURCE_TYPES = {
    "Video",
    "Webinar",
    "Report",
    "Guide",
    "Resource",
    "FAQ",
    "Customer Story",
    "Stories Topic Hub",
}

BLOG_SOURCE_TYPES = {
    "Blog Post",
    "Proposed PILLAR Blog",
    "Blog Post (Series)",
    "Blog Post (Infographic)",
    "Blog Series Hub",
}

BANNED_TERMS = {"best", "compare"}

STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "by",
    "for",
    "from",
    "how",
    "in",
    "into",
    "is",
    "it",
    "its",
    "of",
    "on",
    "or",
    "our",
    "out",
    "that",
    "the",
    "their",
    "they",
    "this",
    "to",
    "with",
    "you",
    "your",
    "omnipress",
    "association",
    "associations",
    "member",
    "members",
}


def stable_hash_int(text: str) -> int:
    return int(hashlib.md5(text.encode("utf-8")).hexdigest(), 16)


def normalize_url(url: Optional[str]) -> Optional[str]:
    if url is None or (isinstance(url, float) and math.isnan(url)):
        return None
    s = str(url).strip()
    if not s or s.lower() == "nan":
        return None
    s = s.split("#")[0].split("?")[0].strip()
    if not s:
        return None
    if not s.endswith("/"):
        s += "/"
    return s


def safe_str(v: Optional[str]) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ""
    s = str(v)
    if s.lower() == "nan":
        return ""
    return s


def clean_anchor(text: str) -> str:
    s = safe_str(text).strip()
    s = s.strip('"').strip("'").strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[\s,]+$", "", s)
    return s


def lead_token(text: str) -> str:
    s = clean_anchor(text).lower()
    return s.split(" ", 1)[0] if s else ""


def contains_banned(anchor: str) -> bool:
    a = anchor.lower()
    return any(re.search(rf"\b{re.escape(t)}\b", a) for t in BANNED_TERMS)


def allow_services_solutions(primary_keyword: str) -> bool:
    k = safe_str(primary_keyword).lower()
    return bool(re.search(r"\b(services|solutions)\b", k))


def filter_anchor(anchor: str, primary_keyword: str) -> str:
    s = clean_anchor(anchor)
    if not s:
        return ""
    if contains_banned(s):
        return ""
    if re.search(r"\b(services|solutions)\b", s.lower()) and not allow_services_solutions(primary_keyword):
        return ""
    return s


def slug_tokens(url: str) -> Set[str]:
    try:
        path = re.sub(r"^https?://", "", url)
        path = path.split("/", 1)[1] if "/" in path else ""
    except Exception:
        path = url
    raw = re.sub(r"[^a-zA-Z0-9]+", " ", path.lower()).strip()
    return {t for t in raw.split() if t and t not in STOPWORDS}


def text_tokens(text: str) -> Set[str]:
    raw = re.sub(r"[^a-zA-Z0-9]+", " ", safe_str(text).lower()).strip()
    return {t for t in raw.split() if t and t not in STOPWORDS}


def build_tokens(row: pd.Series) -> Set[str]:
    toks: Set[str] = set()
    toks |= text_tokens(row.get("Primary Target Keyword", ""))
    toks |= text_tokens(row.get("Secondary Keywords (Expansion Keywords)", ""))
    toks |= text_tokens(row.get("Page Title (or H1)", ""))
    url = normalize_url(row.get("URL")) or ""
    if url:
        toks |= slug_tokens(url)
    toks -= BANNED_TERMS
    return toks


def jaccard(a: Set[str], b: Set[str]) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    if inter == 0:
        return 0.0
    union = len(a | b)
    return inter / union if union else 0.0


def parse_secondary_keywords(raw: str) -> List[str]:
    s = safe_str(raw).strip()
    if not s:
        return []
    parts = [p.strip() for p in re.split(r"[\n,;]+", s) if p.strip()]
    seen = set()
    out: List[str] = []
    for p in parts:
        key = p.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


def short_topic_from_keyword(primary_keyword: str) -> str:
    k = clean_anchor(primary_keyword)
    if not k:
        return ""
    k2 = re.sub(r"\b(software|platform|system|services|service|solutions|solution)\b", "", k, flags=re.I)
    k2 = re.sub(r"\s+", " ", k2).strip()
    return k2 or k


def strip_guide_phrase(text: str) -> str:
    s = clean_anchor(text)
    if not s:
        return ""
    s = re.sub(r"\b(buyer'?s\s+guide|buyers\s+guide|buyer\s+guide)\b", "", s, flags=re.I)
    s = re.sub(r"\bguide\b", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -|")
    return s


def choose_variant(url: str, variants: List[str]) -> str:
    variants = [v for v in variants if v]
    if not variants:
        return ""
    idx = stable_hash_int(url) % len(variants)
    return variants[idx]


def derive_keyword_from_title(title: str) -> str:
    t = clean_anchor(title)
    if not t:
        return ""
    return t.strip("?.!")


def extract_client_name(title: str, url: str) -> str:
    t = safe_str(title).strip()
    t = re.sub(r"^(customer story|case study)\s*[:\-]\s*", "", t, flags=re.I).strip()
    if "Omnipress" in t:
        t2 = re.sub(r"\bOmnipress\b", "", t, flags=re.I).strip(" -|")
        if t2:
            t = t2
    for sep in ["|", "—", "–"]:
        if sep in t:
            part = t.split(sep)[0].strip()
            if len(part) >= 3:
                t = part
                break
    if " - " in t:
        part = t.split(" - ")[0].strip()
        if len(part) >= 3:
            t = part
    if t and len(t.split()) <= 18:
        return t
    u = normalize_url(url) or ""
    slug = u.strip("/").split("/")[-1]
    slug = slug.replace("-", " ")
    slug = re.sub(r"\b(customers?)\b", "", slug, flags=re.I).strip()
    return slug.title() if slug else ""


def case_study_topic(primary_keyword: str, pillar_cluster: str) -> str:
    k = clean_anchor(primary_keyword)
    k = re.sub(r"\b(case study|customer story)\b", "", k, flags=re.I).strip()
    k = re.sub(r"\s+", " ", k).strip(" -|")
    k = k.replace("&", "and")
    if k:
        return k
    pc = safe_str(pillar_cluster).lower()
    if "abstract" in pc:
        return "abstract submissions"
    if "print" in pc or "fulfillment" in pc:
        return "print fulfillment"
    if "conference" in pc:
        return "conference printing"
    if "digital" in pc:
        return "virtual events"
    if "publication" in pc:
        return "publication printing"
    if "training" in pc:
        return "training materials"
    return "their process"


def build_info_anchor(
    url: str,
    primary_keyword: str,
    page_title: str,
    page_type: str,
    is_core_target: bool,
    pillar_cluster: str,
) -> str:
    pk = clean_anchor(primary_keyword)
    title = clean_anchor(page_title)

    topic = short_topic_from_keyword(pk) or short_topic_from_keyword(title)
    topic = strip_guide_phrase(topic) or topic

    if not topic:
        return ""

    pc = safe_str(pillar_cluster).lower()

    # Non-core content: editorial, varied, and mostly not \"how ... works\".
    if not is_core_target:
        variants = [
            f"tips for {topic}",
            f"what to know about {topic}",
            f"a guide to {topic}",
            f"getting started with {topic}",
            f"{topic} workflow tips",
        ]
        if re.search(r"\bchecklist\b", title, flags=re.I) or re.search(r"\bchecklist\b", pk, flags=re.I):
            variants = [f"{topic} checklist", f"checklist for {topic}", f"steps for {topic}"] + variants
        if re.search(r"\binstructions\b", pk, flags=re.I):
            variants = [f"writing better {topic}", f"tips for {topic}"] + variants
        if re.search(r"\btesting\b", pk, flags=re.I):
            variants = [f"testing {topic}", f"how to test {topic}"] + variants
        if "abstract" in pc:
            variants = [f"streamlining {topic}", f"managing {topic}"] + variants
        if "print" in pc or "fulfillment" in pc:
            variants = [f"planning {topic}", f"{topic} workflow"] + variants
        return choose_variant(url, variants)

    # Core/conversion targets: informational but varied lead-ins.
    variants = [
        f"{topic} workflow overview",
        f"{topic} process overview",
        f"understanding {topic}",
        f"streamlining {topic}",
        f"how {topic} works",
    ]
    if "abstract" in pc:
        variants = [
            "abstract submission and review workflow",
            "managing conference abstract reviews",
            f"{topic} workflow overview",
            f"streamlining {topic}",
            f"how {topic} works",
        ]
    if "print" in pc or "fulfillment" in pc:
        variants = [
            "print fulfillment workflow",
            "shipping and fulfillment process",
            f"{topic} process overview",
            f"streamlining {topic}",
        ] + variants
    if "publication" in pc:
        variants = [
            "member publication workflow overview",
            "publishing and distribution workflow",
            f"{topic} process overview",
        ] + variants
    return choose_variant(url, variants)


def build_branded_anchor(
    url: str,
    primary_keyword: str,
    page_title: str,
    page_type: str,
    is_core_target: bool,
    pillar_cluster: str,
) -> str:
    pk = clean_anchor(primary_keyword)
    title = clean_anchor(page_title)
    topic = short_topic_from_keyword(pk) or short_topic_from_keyword(title)
    topic = strip_guide_phrase(topic) or topic

    if not topic and not title:
        return ""

    # Buyer guide / guide: branded but not \"how\".
    if re.search(r"\b(buyer'?s\s+guide|buyers\s+guide|buyer\s+guide)\b", pk, flags=re.I) or re.search(
        r"\b(buyer'?s\s+guide|buyers\s+guide|buyer\s+guide)\b", title, flags=re.I
    ):
        base = strip_guide_phrase(pk) or strip_guide_phrase(title) or topic
        return choose_variant(
            url,
            [
                f"read Omnipress' buyer's guide to {base}",
                f"Omnipress buyer's guide: {base}",
                f"download the Omnipress buyer's guide on {base}",
            ],
        )

    if re.search(r"\bguide\b", pk, flags=re.I) or re.search(r"\bguide\b", title, flags=re.I):
        base = strip_guide_phrase(pk) or strip_guide_phrase(title) or topic
        return choose_variant(
            url,
            [
                f"read Omnipress' guide to {base}",
                f"Omnipress guide: {base}",
                f"from Omnipress: {base} guide",
            ],
        )

    if page_type in {"Video", "Webinar"}:
        topic2 = topic or title
        variants = [
            f"watch the demo: {title}" if title else f"watch the {topic2} demo",
            f"video walkthrough: {topic2}",
            f"on-demand webinar: {topic2}" if page_type == "Webinar" else f"watch the {topic2} walkthrough",
            f"watch the recording: {title}" if title else f"watch {topic2} in action",
        ]
        return choose_variant(url, variants)

    if not is_core_target:
        return choose_variant(
            url,
            [
                f"Omnipress insights on {topic}",
                f"Omnipress on {topic}",
                f"Omnipress resources for {topic}",
            ],
        )

    # Core/conversion pages.
    softwareish = bool(re.search(r"\b(software|platform|system)\b", pk.lower()))
    if softwareish:
        return choose_variant(
            url,
            [
                f"Omnipress {topic} platform",
                f"our {topic} platform",
                f"Omnipress tools for {topic}",
                f"Omnipress {topic} software",
            ],
        )

    # Services: avoid forcing \"services\" unless primary kw includes it.
    if allow_services_solutions(pk):
        return choose_variant(
            url,
            [
                f"Omnipress {pk}",
                f"Omnipress support for {topic}",
                f"work with Omnipress for {topic}",
            ],
        )
    return choose_variant(
        url,
        [
            f"Omnipress support for {topic}",
            f"work with Omnipress for {topic}",
            f"Omnipress {topic}",
        ],
    )


def generate_anchor_options(row: pd.Series, is_core_target: bool) -> List[str]:
    url = normalize_url(row.get("URL")) or ""
    page_type = safe_str(row.get("Page Type (Detailed)", "")).strip()
    title = safe_str(row.get("Page Title (or H1)", "")).strip()
    primary_kw = safe_str(row.get("Primary Target Keyword", "")).strip()
    pillar_cluster = safe_str(row.get("Pillar Cluster", "")).strip()
    secondary = parse_secondary_keywords(row.get("Secondary Keywords (Expansion Keywords)", ""))

    opt1 = clean_anchor(primary_kw) if primary_kw else derive_keyword_from_title(title)

    opt2 = secondary[0] if len(secondary) > 0 else ""
    opt3 = secondary[1] if len(secondary) > 1 else ""

    if not opt2 and primary_kw:
        topic = short_topic_from_keyword(primary_kw)
        if topic and topic.lower() != primary_kw.lower():
            if re.search(r"\bsoftware\b", primary_kw, flags=re.I):
                opt2 = f"software for {topic}"
            elif re.search(r"\bplatform\b", primary_kw, flags=re.I):
                opt2 = f"{topic} platform"
            else:
                opt2 = topic

    if not opt3 and primary_kw:
        if re.search(r"\bsoftware\b", primary_kw, flags=re.I):
            opt3 = primary_kw.replace("software", "platform")
        elif re.search(r"\bplatform\b", primary_kw, flags=re.I):
            opt3 = primary_kw.replace("platform", "software")

    opt4 = build_info_anchor(url, primary_kw, title, page_type, is_core_target, pillar_cluster)

    if page_type == "Customer Story":
        client = extract_client_name(title, safe_str(row.get("URL")))
        topic = case_study_topic(primary_kw, pillar_cluster)
        opt2 = opt2 or (f"read {client}'s story" if client else "read their story")
        opt3 = f"{client} streamlined {topic}" if client else f"streamlining {topic}"
        opt4 = opt4 or (f"results from {client}'s {topic} process" if client else f"results from their {topic} process")
        opt5 = f"{client}'s experience with Omnipress" if client else "their experience with Omnipress"
    else:
        opt5 = build_branded_anchor(url, primary_kw, title, page_type, is_core_target, pillar_cluster)

    opts = [opt1, opt2, opt3, opt4, opt5]

    # Reduce repetitive opening tokens across options (e.g., "how", "see").
    # Keep exact match untouched; only rewrite informational/branded variants.
    mutable_variants = {
        3: [
            "workflow for {topic}",
            "practical guidance on {topic}",
            "key considerations for {topic}",
        ],
        4: [
            "Omnipress perspective on {topic}",
            "Omnipress resources for {topic}",
            "inside Omnipress: {topic}",
        ],
    }
    topic_for_rewrite = strip_guide_phrase(short_topic_from_keyword(primary_kw) or short_topic_from_keyword(title))
    if topic_for_rewrite:
        seen_leads = set()
        for i, raw in enumerate(opts):
            token = lead_token(raw)
            if not token:
                continue
            if i == 0:
                seen_leads.add(token)
                continue
            if token in seen_leads and i in mutable_variants:
                repls = [v.format(topic=topic_for_rewrite) for v in mutable_variants[i]]
                opts[i] = choose_variant(url + f"-opt{i+1}", repls)
                token = lead_token(opts[i])
            if token:
                seen_leads.add(token)

    cleaned: List[str] = []
    seen = set()
    for o in opts:
        o2 = filter_anchor(o, primary_kw)
        if not o2:
            cleaned.append("")
            continue
        key = o2.lower()
        if key in seen:
            cleaned.append("")
            continue
        seen.add(key)
        cleaned.append(o2)

    # Ensure at least 2 anchors
    if sum(1 for x in cleaned if x) < 2:
        topic = strip_guide_phrase(short_topic_from_keyword(primary_kw) or short_topic_from_keyword(title)) or clean_anchor(
            primary_kw
        )
        for fb in [f"tips for {topic}", f"Omnipress on {topic}"]:
            fb2 = filter_anchor(fb, primary_kw)
            if fb2 and fb2.lower() not in seen:
                for i in range(5):
                    if not cleaned[i]:
                        cleaned[i] = fb2
                        seen.add(fb2.lower())
                        break
            if sum(1 for x in cleaned if x) >= 2:
                break

    return cleaned


def best_candidate_with_score(source_tokens: Set[str], candidates: pd.DataFrame) -> Tuple[Optional[pd.Series], float]:
    if candidates.empty:
        return None, -1.0
    scored: List[Tuple[float, str, int]] = []
    for idx, r in candidates.iterrows():
        score = jaccard(source_tokens, r["__tokens"])
        url = normalize_url(r["URL"]) or safe_str(r["URL"])
        scored.append((score, url, idx))
    scored.sort(key=lambda x: (-x[0], x[1]))
    score, _, idx = scored[0]
    return candidates.loc[idx], score


def allocate_core_mix_counts(n: int) -> Tuple[int, int, int, int]:
    if n <= 0:
        return (0, 0, 0, 0)

    bounds = {
        "exact": (math.floor(n * 0.15), math.ceil(n * 0.20)),
        "partial": (math.floor(n * 0.40), math.ceil(n * 0.50)),
        "info": (math.floor(n * 0.20), math.ceil(n * 0.30)),
        "branded": (math.floor(n * 0.10), math.ceil(n * 0.20)),
    }

    target = {"exact": 0.18, "partial": 0.45, "info": 0.25, "branded": 0.12}
    combos: List[Tuple[float, int, int, int, int]] = []

    for exact in range(bounds["exact"][0], bounds["exact"][1] + 1):
        for partial in range(bounds["partial"][0], bounds["partial"][1] + 1):
            for info in range(bounds["info"][0], bounds["info"][1] + 1):
                branded = n - exact - partial - info
                if branded < 0:
                    continue
                if branded < bounds["branded"][0] or branded > bounds["branded"][1]:
                    continue
                err = 0.0
                for k, v in [("exact", exact), ("partial", partial), ("info", info), ("branded", branded)]:
                    err += (v / n - target[k]) ** 2
                combos.append((err, exact, partial, info, branded))

    if not combos:
        exact = 1 if n >= 5 else 0
        info = 1 if n >= 4 else 0
        branded = 1 if n >= 6 else 0
        partial = max(0, n - exact - info - branded)
        return exact, partial, info, branded

    combos.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
    _, exact, partial, info, branded = combos[0]
    return exact, partial, info, branded


def evenly_spaced_positions(n: int, count: int, offset: float, taken: Set[int]) -> List[int]:
    if count <= 0 or n <= 0:
        return []
    positions = []
    step = n / count
    start = step / 2 + offset
    for i in range(count):
        pos = int(math.floor(start + i * step))
        pos = max(0, min(n - 1, pos))
        if pos in taken:
            for j in range(n):
                cand = (pos + j) % n
                if cand not in taken:
                    pos = cand
                    break
        if pos in taken:
            continue
        taken.add(pos)
        positions.append(pos)
    return positions


def assign_anchor_categories_for_target(n: int, is_core_target: bool) -> List[str]:
    if n <= 0:
        return []
    if is_core_target:
        exact, _, info, branded = allocate_core_mix_counts(n)
        cats = ["partial"] * n
        taken: Set[int] = set()
        for pos in evenly_spaced_positions(n, exact, 0.0, taken):
            cats[pos] = "exact"
        for pos in evenly_spaced_positions(n, info, 0.17, taken):
            cats[pos] = "info"
        for pos in evenly_spaced_positions(n, branded, 0.34, taken):
            cats[pos] = "branded"
        return cats
    pattern = ["info", "partial2", "branded", "partial3"]
    return [pattern[i % len(pattern)] for i in range(n)]


def pick_anchor_for_category(options: List[str], category: str, partial_toggle: int) -> Tuple[str, int, int]:
    opt1, opt2, opt3, opt4, opt5 = options

    if category == "exact":
        if opt1:
            return opt1, 1, partial_toggle
        for i, opt in enumerate([opt2, opt3, opt4, opt5], start=2):
            if opt:
                return opt, i, partial_toggle
        return "", 0, partial_toggle

    if category == "info":
        if opt4:
            return opt4, 4, partial_toggle
        for i, opt in [(2, opt2), (3, opt3), (5, opt5), (1, opt1)]:
            if opt:
                return opt, i, partial_toggle
        return "", 0, partial_toggle

    if category == "branded":
        if opt5:
            return opt5, 5, partial_toggle
        for i, opt in [(4, opt4), (2, opt2), (3, opt3), (1, opt1)]:
            if opt:
                return opt, i, partial_toggle
        return "", 0, partial_toggle

    if category in {"partial", "partial2", "partial3"}:
        preferred = [2, 3] if partial_toggle % 2 == 0 else [3, 2]
        for idx in preferred:
            opt = opt2 if idx == 2 else opt3
            if opt:
                return opt, idx, partial_toggle + 1
        for i, opt in [(4, opt4), (5, opt5), (1, opt1)]:
            if opt:
                return opt, i, partial_toggle
        return "", 0, partial_toggle

    for i, opt in [(2, opt2), (3, opt3), (4, opt4), (5, opt5), (1, opt1)]:
        if opt:
            return opt, i, partial_toggle
    return "", 0, partial_toggle


def category_label(is_core: bool, opt_num: int) -> str:
    if is_core:
        if opt_num == 1:
            return "Exact (Opt 1 | 15–20%)"
        if opt_num == 2:
            return "Partial/Phrase (Opt 2 | 40–50%)"
        if opt_num == 3:
            return "Partial/Phrase (Opt 3 | 40–50%)"
        if opt_num == 4:
            return "Informational/Narrative (Opt 4 | 20–30%)"
        if opt_num == 5:
            return "Branded/Narrative (Opt 5 | 10–20%)"
    else:
        if opt_num == 1:
            return "Exact (Opt 1 | Optional | not enforced)"
        if opt_num in {2, 3}:
            return "Partial/Phrase (Preferred | non-core)"
        if opt_num == 4:
            return "Informational/Narrative (Preferred | non-core)"
        if opt_num == 5:
            return "Branded/Narrative (Preferred | non-core)"
    return ""


def write_df(ws, df: pd.DataFrame) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=1, column=c_idx)
                cell.font = header_font
                cell.fill = header_fill
    ws.freeze_panes = "A2"


def main() -> None:
    pm = pd.read_csv(PAGE_MASTER_PATH)
    core = pd.read_csv(CORE_CLUSTERS_PATH)

    pm["URL_norm"] = pm["URL"].apply(normalize_url)
    pm["Pillar_norm"] = pm["Pillar URL"].apply(normalize_url)
    core["URL_norm"] = core["URL"].apply(normalize_url)

    pillar_set = set(core["URL_norm"].dropna())
    pillar_cluster_map = dict(zip(core["URL_norm"], core["Topic Cluster"]))

    in_scope_supporting = pm[pm["Pillar_norm"].isin(pillar_set)].copy()
    pillar_pages = pm[pm["URL_norm"].isin(pillar_set)].copy()

    missing_pillars = sorted(pillar_set - set(pillar_pages["URL_norm"].dropna()))
    if missing_pillars:
        add_rows = []
        for u in missing_pillars:
            r = core[core["URL_norm"] == u].iloc[0].to_dict()
            add_rows.append(
                {
                    "URL": u,
                    "URL_norm": u,
                    "Page Title (or H1)": "",
                    "Page Type": "",
                    "Page Type (Detailed)": r.get("Page Type (Detailed)", ""),
                    "Pillar URL": None,
                    "Pillar_norm": None,
                    "Topic Cluster": r.get("Topic Cluster", ""),
                    "Cluster ID": r.get("CLUSTER ID", ""),
                    "Primary Target Keyword": r.get("Primary Target Keyword", ""),
                    "Secondary Keywords (Expansion Keywords)": "",
                    "Search Intent": "",
                    "Performance Snapshot": "",
                    "Suggested Action": "",
                    "Notes": "Added from core cluster file; missing in page master.",
                }
            )
        pillar_pages = pd.concat([pillar_pages, pd.DataFrame(add_rows)], ignore_index=True)

    in_scope_supporting["Pillar Cluster"] = in_scope_supporting["Pillar_norm"].map(pillar_cluster_map)
    in_scope_supporting["Pillar URL (Normalized)"] = in_scope_supporting["Pillar_norm"]

    pillar_pages["Pillar Cluster"] = pillar_pages["URL_norm"].map(pillar_cluster_map)
    pillar_pages["Pillar URL (Normalized)"] = pillar_pages["URL_norm"]

    pages_all = pd.concat([in_scope_supporting, pillar_pages], ignore_index=True)
    pages_all = pages_all.sort_values(by=["URL_norm", "Pillar_norm"], na_position="last")
    pages_all = pages_all.drop_duplicates(subset=["URL_norm"], keep="first").copy()

    pages_all["type"] = pages_all["Page Type (Detailed)"].fillna("")
    pages_all["__tokens"] = pages_all.apply(build_tokens, axis=1)

    pages_by_url: Dict[str, pd.Series] = {u: r for u, r in pages_all.set_index("URL_norm").iterrows()}

    # Build link tasks
    sources = in_scope_supporting.copy()
    sources["type"] = sources["Page Type (Detailed)"].fillna("")
    sources["is_asset"] = sources["type"].isin(ASSET_SOURCE_TYPES)
    sources["is_blog"] = sources["type"].isin(BLOG_SOURCE_TYPES)
    sources["is_conv"] = sources["type"].isin(CORE_TARGET_TYPES)

    blog_conv_sources = sources[(sources["is_blog"] | sources["is_conv"]) & (~sources["is_asset"])].copy()
    asset_sources = sources[sources["is_asset"]].copy()

    def cluster_pages(pillar_cluster: str) -> pd.DataFrame:
        return pages_all[pages_all["Pillar Cluster"] == pillar_cluster].copy()

    def get_tokens_for_url(url: str) -> Set[str]:
        r = pages_by_url.get(url)
        return r["__tokens"] if r is not None else set()

    link_tasks: List[Dict[str, object]] = []

    def add_task(
        pillar_cluster: str,
        pillar_url: str,
        source_url: str,
        source_type: str,
        target_url: str,
        link_tier: str,
    ) -> None:
        link_tasks.append(
            {
                "Pillar Cluster": pillar_cluster,
                "Pillar URL": pillar_url,
                "Source URL": source_url,
                "Source Type (Detailed)": source_type,
                "Target URL": target_url,
                "Link Tier": link_tier,
            }
        )

    for _, src in blog_conv_sources.iterrows():
        source_url = normalize_url(src["URL"])
        pillar_url = normalize_url(src["Pillar URL"])
        pillar_cluster = pillar_cluster_map.get(pillar_url)
        if not source_url or not pillar_url or not pillar_cluster:
            continue

        source_type = safe_str(src["Page Type (Detailed)"]).strip()
        if source_url != pillar_url:
            add_task(pillar_cluster, pillar_url, source_url, source_type, pillar_url, "supporting_to_pillar")

        cps = cluster_pages(pillar_cluster)
        exclude = {source_url, pillar_url}
        src_tokens = get_tokens_for_url(source_url)

        # Cross-link 1: conversion/profit
        conv_candidates = cps[
            (cps["type"].isin(CORE_TARGET_TYPES))
            & (~cps["URL_norm"].isin(exclude))
            & (cps["URL_norm"] != pillar_url)
        ]
        conv_pick, _ = best_candidate_with_score(src_tokens, conv_candidates)
        conv_url = normalize_url(conv_pick["URL"]) if conv_pick is not None else None
        if not conv_url:
            any_pick, _ = best_candidate_with_score(src_tokens, cps[(~cps["URL_norm"].isin(exclude))])
            conv_url = normalize_url(any_pick["URL"]) if any_pick is not None else None
        if conv_url and conv_url not in exclude:
            add_task(pillar_cluster, pillar_url, source_url, source_type, conv_url, "supporting_to_supporting")
            exclude.add(conv_url)

        # Cross-link 2: educational/proof (prefer assets and customer stories if reasonably relevant)
        edu_assets = cps[(cps["type"].isin(ASSET_SOURCE_TYPES)) & (~cps["URL_norm"].isin(exclude))]
        edu_blogs = cps[(cps["type"].isin(BLOG_SOURCE_TYPES)) & (~cps["URL_norm"].isin(exclude))]

        best_cs, score_cs = best_candidate_with_score(src_tokens, edu_assets[edu_assets["type"] == "Customer Story"])
        best_asset, score_asset = best_candidate_with_score(src_tokens, edu_assets[edu_assets["type"] != "Customer Story"])
        best_blog, score_blog = best_candidate_with_score(src_tokens, edu_blogs)

        edu_pick = None
        if best_cs is not None and score_cs >= max(0.05, score_blog - 0.15):
            edu_pick = best_cs
        elif best_asset is not None and score_asset >= max(0.05, score_blog - 0.08):
            edu_pick = best_asset
        elif best_blog is not None:
            edu_pick = best_blog
        else:
            any_pick, _ = best_candidate_with_score(src_tokens, cps[(~cps["URL_norm"].isin(exclude))])
            edu_pick = any_pick

        edu_url = normalize_url(edu_pick["URL"]) if edu_pick is not None else None
        if edu_url and edu_url not in exclude:
            add_task(pillar_cluster, pillar_url, source_url, source_type, edu_url, "supporting_to_supporting")

    # Assets: 1 link only
    for _, src in asset_sources.iterrows():
        source_url = normalize_url(src["URL"])
        pillar_url = normalize_url(src["Pillar URL"])
        pillar_cluster = pillar_cluster_map.get(pillar_url)
        if not source_url or not pillar_url or not pillar_cluster:
            continue
        source_type = safe_str(src["Page Type (Detailed)"]).strip()
        cps = cluster_pages(pillar_cluster)
        exclude = {source_url, pillar_url}
        conv_candidates = cps[
            (cps["type"].isin(CORE_TARGET_TYPES))
            & (~cps["URL_norm"].isin(exclude))
            & (cps["URL_norm"] != pillar_url)
        ]
        src_tokens = get_tokens_for_url(source_url)
        conv_pick, _ = best_candidate_with_score(src_tokens, conv_candidates)
        target_url = normalize_url(conv_pick["URL"]) if conv_pick is not None else None
        if not target_url:
            target_url = pillar_url
            link_tier = "asset_to_pillar"
        else:
            link_tier = "asset_to_conversion"
        if target_url and target_url != source_url:
            add_task(pillar_cluster, pillar_url, source_url, source_type, target_url, link_tier)

    link_df = pd.DataFrame(link_tasks)
    link_df["Source URL"] = link_df["Source URL"].apply(normalize_url)
    link_df["Target URL"] = link_df["Target URL"].apply(normalize_url)
    link_df = link_df[(link_df["Source URL"].notna()) & (link_df["Target URL"].notna())]
    link_df = link_df[link_df["Source URL"] != link_df["Target URL"]]
    link_df = link_df.drop_duplicates(subset=["Source URL", "Target URL"], keep="first").copy()
    link_df["__tier_sort"] = link_df["Link Tier"].apply(lambda t: 0 if t == "supporting_to_pillar" else 1)
    link_df = link_df.sort_values(by=["Source URL", "__tier_sort", "Target URL"]).drop(columns=["__tier_sort"]).reset_index(drop=True)

    # Anchor options for ALL in-scope pages
    anchor_rows = []
    for _, r in pages_all.iterrows():
        url = normalize_url(r.get("URL"))
        if not url:
            continue
        page_type = safe_str(r.get("Page Type (Detailed)", "")).strip()
        is_core_target = (page_type in CORE_TARGET_TYPES) or (url in pillar_set)
        opts = generate_anchor_options(r, is_core_target)
        anchor_rows.append(
            {
                "Pillar Cluster": safe_str(r.get("Pillar Cluster", "")).strip(),
                "Pillar URL": normalize_url(r.get("Pillar URL (Normalized)")) or "",
                "Subcluster": safe_str(r.get("Topic Cluster", "")).strip(),
                "URL": url,
                "Page Title (or H1)": safe_str(r.get("Page Title (or H1)", "")).strip(),
                "Page Type (Detailed)": page_type,
                "Search Intent": safe_str(r.get("Search Intent", "")).strip(),
                "Primary Target Keyword": safe_str(r.get("Primary Target Keyword", "")).strip(),
                "Anchor Option 1 (Exact)": opts[0],
                "Anchor Option 2": opts[1],
                "Anchor Option 3": opts[2],
                "Anchor Option 4": opts[3],
                "Anchor Option 5": opts[4],
                "Target Is Core/Conversion": is_core_target,
            }
        )
    anchor_df = pd.DataFrame(anchor_rows)

    # Assign anchors to tasks
    options_map: Dict[str, List[str]] = {}
    core_target_map: Dict[str, bool] = {}
    for _, r in anchor_df.iterrows():
        options_map[r["URL"]] = [
            safe_str(r["Anchor Option 1 (Exact)"]).strip(),
            safe_str(r["Anchor Option 2"]).strip(),
            safe_str(r["Anchor Option 3"]).strip(),
            safe_str(r["Anchor Option 4"]).strip(),
            safe_str(r["Anchor Option 5"]).strip(),
        ]
        core_target_map[r["URL"]] = bool(r["Target Is Core/Conversion"])

    assigned_rows = []
    for target_url, group in link_df.groupby("Target URL", sort=True):
        group = group.sort_values(by=["Source URL", "Link Tier", "Target URL"]).copy()
        n = len(group)
        is_core_target = core_target_map.get(target_url, False)
        cats = assign_anchor_categories_for_target(n, is_core_target)
        partial_toggle = 0
        for (_, row_task), cat in zip(group.iterrows(), cats):
            opts = options_map.get(target_url, ["", "", "", "", ""])
            anchor_text, opt_used, partial_toggle = pick_anchor_for_category(opts, cat, partial_toggle)
            assigned_rows.append(
                {
                    **row_task.to_dict(),
                    "Target Is Core/Conversion": is_core_target,
                    "Anchor Category": cat,
                    "Anchor Option Used": opt_used,
                    "Anchor Text": anchor_text,
                }
            )
    assigned_df = pd.DataFrame(assigned_rows)

    # Enrich with target + source metadata
    anchor_meta = anchor_df[
        [
            "URL",
            "Page Type (Detailed)",
            "Primary Target Keyword",
            "Anchor Option 1 (Exact)",
            "Anchor Option 2",
            "Anchor Option 3",
            "Anchor Option 4",
            "Anchor Option 5",
        ]
    ].rename(
        columns={
            "URL": "Target URL",
            "Page Type (Detailed)": "Target Type (Detailed)",
            "Primary Target Keyword": "Target Primary Keyword",
        }
    )
    assigned_df = assigned_df.merge(anchor_meta, on="Target URL", how="left")
    source_meta = anchor_df[["URL", "Page Type (Detailed)", "Primary Target Keyword"]].rename(
        columns={
            "URL": "Source URL",
            "Page Type (Detailed)": "Source Type (Detailed) (From Master)",
            "Primary Target Keyword": "Source Primary Keyword",
        }
    )
    assigned_df = assigned_df.merge(source_meta, on="Source URL", how="left")
    assigned_df["__tier_sort"] = assigned_df["Link Tier"].apply(lambda t: 0 if t == "supporting_to_pillar" else 1)
    assigned_df = assigned_df.sort_values(by=["Source URL", "__tier_sort", "Target URL"]).drop(columns=["__tier_sort"]).reset_index(drop=True)

    # Worksheet CSV export (3 cols)
    worksheet_df = assigned_df[["Source URL", "Target URL", "Anchor Text"]].rename(
        columns={"Source URL": "Content URL", "Target URL": "Suggested Target URL", "Anchor Text": "Suggested Anchor Text"}
    )
    worksheet_df.to_csv(OUT_WORKSHEET_CSV, index=False)

    # Sterling-style Anchor Bank CSVs (long format)
    formatted_rows = []
    for _, r in anchor_df.iterrows():
        url = r["URL"]
        is_core = bool(r["Target Is Core/Conversion"])
        name = clean_anchor(r.get("Page Title (or H1)", "")) or clean_anchor(r.get("Primary Target Keyword", ""))
        opts = [
            r.get("Anchor Option 1 (Exact)", ""),
            r.get("Anchor Option 2", ""),
            r.get("Anchor Option 3", ""),
            r.get("Anchor Option 4", ""),
            r.get("Anchor Option 5", ""),
        ]
        for i, opt in enumerate(opts, start=1):
            optc = clean_anchor(opt)
            if not optc:
                continue
            formatted_rows.append(
                {
                    "Core Page Name": name,
                    "Core Page URL": url,
                    "Category": category_label(is_core, i),
                    "Anchor Text": optc,
                    "Target Is Core/Conversion": is_core,
                }
            )
    formatted_df = pd.DataFrame(formatted_rows)
    all_targets_csv_df = formatted_df[["Core Page Name", "Core Page URL", "Category", "Anchor Text"]].copy()
    core_targets_csv_df = formatted_df[formatted_df["Target Is Core/Conversion"]][
        ["Core Page Name", "Core Page URL", "Category", "Anchor Text"]
    ].copy()
    all_targets_csv_df.to_csv(OUT_ANCHOR_BANK_ALL, index=False)
    core_targets_csv_df.to_csv(OUT_ANCHOR_BANK_CORE, index=False)

    # QA (counts + anchor scans + mix)
    qa_summary = {
        "total_sources_in_scope": int(sources["URL_norm"].nunique()),
        "blog_conv_sources": int(blog_conv_sources["URL_norm"].nunique()),
        "asset_sources": int(asset_sources["URL_norm"].nunique()),
        "total_link_tasks": int(len(link_df)),
    }

    counts = assigned_df.groupby("Source URL").size().rename("task_count").reset_index()
    counts = counts.merge(
        sources[["URL_norm", "type", "Pillar Cluster", "Pillar_norm", "is_asset"]].rename(
            columns={"URL_norm": "Source URL", "type": "Source Type", "Pillar_norm": "Pillar URL"}
        ),
        on="Source URL",
        how="left",
    )
    cluster_supporting_counts = sources.groupby("Pillar Cluster")["URL_norm"].nunique().to_dict()
    expected_by_rule = []
    max_possible = []
    status = []
    exception_reason = []
    for _, r in counts.iterrows():
        is_asset = bool(r["is_asset"])
        if is_asset:
            expected = 1
            max_p = 1
            st = "PASS" if r["task_count"] == 1 else "FAIL"
            reason = ""
        else:
            expected = 3
            cluster = safe_str(r["Pillar Cluster"])
            cluster_size = int(cluster_supporting_counts.get(cluster, 0))
            available_targets = max(0, cluster_size - 1)
            max_cross = min(2, available_targets)
            max_p = 1 + max_cross
            if r["task_count"] == expected:
                st = "PASS"
                reason = ""
            elif r["task_count"] == max_p and max_p < expected:
                st = "EXCEPTION"
                reason = f"Only {available_targets} other in-scope supporting page(s) in cluster; cannot create 2 cross-links."
            else:
                st = "FAIL"
                reason = ""
        expected_by_rule.append(expected)
        max_possible.append(max_p)
        status.append(st)
        exception_reason.append(reason)
    counts["expected_by_rule"] = expected_by_rule
    counts["max_possible_in_cluster"] = max_possible
    counts["status"] = status
    counts["exception_reason"] = exception_reason
    qa_summary["source_failures"] = int((counts["status"] == "FAIL").sum())
    qa_summary["source_exceptions"] = int((counts["status"] == "EXCEPTION").sum())
    qa_summary["duplicate_pairs"] = int(assigned_df.duplicated(subset=["Source URL", "Target URL"]).sum())
    qa_summary["self_links"] = int((assigned_df["Source URL"] == assigned_df["Target URL"]).sum())

    qa_summary["anchors_with_best_compare"] = int(
        assigned_df["Anchor Text"].fillna("").astype(str).str.contains(r"\b(best|compare)\b", case=False, regex=True).sum()
    )

    mask_solserv = assigned_df["Anchor Text"].fillna("").astype(str).str.contains(r"\b(services|solutions)\b", case=False, regex=True)
    viol = []
    for _, r in assigned_df[mask_solserv].iterrows():
        pk = safe_str(r.get("Target Primary Keyword", "")).lower()
        if re.search(r"\b(services|solutions)\b", pk):
            continue
        viol.append(r)
    qa_summary["anchors_solserv_violations"] = len(viol)

    # Customer story narrative QA: require at least one anchor mentioning Omnipress + one narrative verb
    cs_issues = []
    cs_df = anchor_df[anchor_df["Page Type (Detailed)"] == "Customer Story"]
    for _, r in cs_df.iterrows():
        joined = " ".join(
            [safe_str(r.get(c, "")) for c in ["Anchor Option 2", "Anchor Option 3", "Anchor Option 4", "Anchor Option 5"]]
        ).lower()
        if ("omnipress" in joined) and any(v in joined for v in ["read", "streamlined", "results", "experience", "improved"]):
            continue
        cs_issues.append(r["URL"])
    qa_summary["customer_story_anchor_issues"] = len(cs_issues)

    # Video/webinar narrative QA
    vw_issues = []
    vw_df = anchor_df[anchor_df["Page Type (Detailed)"].isin(["Video", "Webinar"])]
    for _, r in vw_df.iterrows():
        joined = " ".join(
            [safe_str(r.get(c, "")) for c in ["Anchor Option 2", "Anchor Option 3", "Anchor Option 4", "Anchor Option 5"]]
        ).lower()
        if any(w in joined for w in ["watch", "demo", "walkthrough", "recording", "webinar", "video"]):
            continue
        vw_issues.append(r["URL"])
    qa_summary["video_webinar_anchor_issues"] = len(vw_issues)

    # Mix by target QA (core targets only)
    def rollup_cat(cat: str) -> str:
        return "partial" if cat.startswith("partial") else cat

    mix_rows = []
    for target_url, grp in assigned_df.groupby("Target URL", sort=True):
        is_core_target = bool(grp["Target Is Core/Conversion"].iloc[0])
        if not is_core_target:
            continue
        n = len(grp)
        vc = grp["Anchor Category"].fillna("").apply(rollup_cat).value_counts().to_dict()
        exact = int(vc.get("exact", 0))
        partial = int(vc.get("partial", 0))
        info = int(vc.get("info", 0))
        branded = int(vc.get("branded", 0))
        bounds = {
            "exact": (math.floor(n * 0.15), math.ceil(n * 0.20)),
            "partial": (math.floor(n * 0.40), math.ceil(n * 0.50)),
            "info": (math.floor(n * 0.20), math.ceil(n * 0.30)),
            "branded": (math.floor(n * 0.10), math.ceil(n * 0.20)),
        }

        def in_bounds(val: int, key: str) -> bool:
            lo, hi = bounds[key]
            return lo <= val <= hi

        qa_pass = all(
            [
                in_bounds(exact, "exact"),
                in_bounds(partial, "partial"),
                in_bounds(info, "info"),
                in_bounds(branded, "branded"),
                exact + partial + info + branded == n,
            ]
        )
        mix_rows.append(
            {
                "Target URL": target_url,
                "Target Is Core/Conversion": True,
                "inbound_links": n,
                "Mix QA": "PASS" if qa_pass else "FAIL",
                "exact": exact,
                "exact_pct": exact / n,
                "partial": partial,
                "partial_pct": partial / n,
                "info": info,
                "info_pct": info / n,
                "branded": branded,
                "branded_pct": branded / n,
                "Target Primary Keyword (From Bank)": safe_str(grp["Target Primary Keyword"].iloc[0]),
            }
        )
    mix_df = pd.DataFrame(mix_rows).sort_values(by=["Mix QA", "Target URL"]).reset_index(drop=True)
    qa_summary["core_targets_mix_failures"] = int((mix_df["Mix QA"] == "FAIL").sum()) if not mix_df.empty else 0

    # Workbook outputs
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "Worksheet (ADD_CONTEXTUAL_UPD)"
    write_df(ws0, worksheet_df)

    ws1 = wb.create_sheet("Link Tasks (Detailed)")
    link_tasks_detailed_cols = [
        "Pillar Cluster",
        "Pillar URL",
        "Source URL",
        "Source Type (Detailed)",
        "Target URL",
        "Target Type (Detailed)",
        "Target Is Core/Conversion",
        "Link Tier",
        "Anchor Category",
        "Anchor Option Used",
        "Anchor Text",
        "Anchor Option 1 (Exact)",
        "Anchor Option 2",
        "Anchor Option 3",
        "Anchor Option 4",
        "Anchor Option 5",
    ]
    write_df(ws1, assigned_df[link_tasks_detailed_cols].copy())

    ws2 = wb.create_sheet("Anchor Text Library")
    anchor_library_cols = [
        "Pillar Cluster",
        "Pillar URL",
        "Subcluster",
        "URL",
        "Page Type (Detailed)",
        "Search Intent",
        "Primary Target Keyword",
        "Anchor Option 1 (Exact)",
        "Anchor Option 2",
        "Anchor Option 3",
        "Anchor Option 4",
        "Anchor Option 5",
    ]
    write_df(ws2, anchor_df[anchor_library_cols].copy())

    ws3 = wb.create_sheet("Anchor Bank (Formatted)")
    write_df(ws3, all_targets_csv_df)

    ws4 = wb.create_sheet("Inputs - Core Pages")
    write_df(ws4, core[["URL", "CLUSTER ID", "Page Type (Detailed)", "Topic Cluster", "Primary Target Keyword"]].copy())

    ws5 = wb.create_sheet("Inputs - Page Master (In-Scope)")
    pm_cols = [
        "URL",
        "Page Title (or H1)",
        "Page Type (Detailed)",
        "Pillar URL",
        "Topic Cluster",
        "Primary Target Keyword",
        "Secondary Keywords (Expansion Keywords)",
        "Search Intent",
        "Notes",
    ]
    write_df(ws5, in_scope_supporting[pm_cols].copy())

    ws6 = wb.create_sheet("QA")
    qa_rows: List[List[object]] = []
    qa_rows.append(["Metric", "Value"])
    for k, v in qa_summary.items():
        qa_rows.append([k, v])
    qa_rows.append([])
    qa_rows.append(
        [
            "Source URL",
            "Source Type",
            "Pillar Cluster",
            "expected_by_rule",
            "max_possible_in_cluster",
            "task_count",
            "status",
            "exception_reason",
        ]
    )
    for _, r in counts.sort_values(by=["status", "Pillar Cluster", "Source URL"]).iterrows():
        qa_rows.append(
            [
                r["Source URL"],
                r["Source Type"],
                r["Pillar Cluster"],
                r["expected_by_rule"],
                r["max_possible_in_cluster"],
                r["task_count"],
                r["status"],
                r["exception_reason"],
            ]
        )
    qa_rows.append([])
    qa_rows.append(["Mix By Target (Core Targets Only)"])
    qa_rows.append(list(mix_df.columns))
    for _, r in mix_df.iterrows():
        qa_rows.append([r[c] for c in mix_df.columns])
    for row in qa_rows:
        ws6.append(row)
    ws6.freeze_panes = "A2"

    wb.save(OUT_WORKBOOK)

    wb2 = openpyxl.Workbook()
    wb2.remove(wb2.active)
    ws_m = wb2.create_sheet("Mix By Target")
    write_df(ws_m, mix_df)
    ws_a = wb2.create_sheet("Assignments (Inbound Links)")
    write_df(ws_a, assigned_df[link_tasks_detailed_cols].copy())
    ws_b = wb2.create_sheet("Anchor Bank (Options)")
    write_df(ws_b, anchor_df[anchor_library_cols].copy())
    ws_f = wb2.create_sheet("Anchor Bank (Formatted)")
    write_df(ws_f, all_targets_csv_df)
    ws_r = wb2.create_sheet("Rules")
    rules_df = pd.DataFrame(
        [
            {
                "Section": "Core/conversion targets (mix enforced per Target URL)",
                "Deterministic assignment": "Spread categories across inbound links via evenly-spaced positions; rotate Partial between Opt 2 and Opt 3.",
                "Mix (goal)": "Exact 15–20%, Partial 40–50%, Informational 20–30%, Branded 10–20%",
                "QA method": "Integer bounds via floor/ceil of % ranges per target URL (handles small n).",
            },
            {
                "Section": "Non-core targets (blogs/assets)",
                "Deterministic assignment": "Narrative rotation: Opt 4 → Opt 2 → Opt 5 → Opt 3 (repeat). Exact not enforced.",
                "Mix (goal)": "Prefer Partial/Info/Branded with narrative phrasing",
                "QA method": "No exact-match enforcement; verify banned terms removed and narrative present for assets.",
            },
            {
                "Section": "Banned modifiers",
                "Deterministic assignment": "Do not generate anchors with best/compare; do not add services/solutions unless in primary keyword.",
                "Mix (goal)": "Natural anchors; avoid spam patterns",
                "QA method": "Regex scan across anchor options and assigned anchors.",
            },
        ]
    )
    write_df(ws_r, rules_df)
    ws_q = wb2.create_sheet("QA")
    qa2_rows: List[List[object]] = []
    qa2_rows.append(["Metric", "Value"])
    for k, v in qa_summary.items():
        qa2_rows.append([k, v])
    qa2_rows.append([])
    qa2_rows.append(["Customer Story anchor issues (count)", qa_summary["customer_story_anchor_issues"]])
    for u in cs_issues[:50]:
        qa2_rows.append([u])
    qa2_rows.append([])
    qa2_rows.append(["Video/Webinar anchor issues (count)", qa_summary["video_webinar_anchor_issues"]])
    for u in vw_issues[:50]:
        qa2_rows.append([u])
    for row in qa2_rows:
        ws_q.append(row)
    ws_q.freeze_panes = "A2"
    wb2.save(OUT_ASSIGNMENTS_XLSX)

    # Console summary for quick iteration
    def starts_with(word: str, series: pd.Series) -> int:
        return int(series.fillna("").astype(str).str.strip().str.match(rf"(?i)^{re.escape(word)}\b").sum())

    anchors_bank = all_targets_csv_df["Anchor Text"]
    anchors_ws = worksheet_df["Suggested Anchor Text"]
    print("v3 anchor bank total anchors:", len(anchors_bank))
    print("v3 anchor bank starts with how:", starts_with("how", anchors_bank))
    print("v3 anchor bank starts with see:", starts_with("see", anchors_bank))
    print("v3 worksheet rows:", len(anchors_ws))
    print("v3 worksheet starts with how:", starts_with("how", anchors_ws))
    print("v3 worksheet starts with see:", starts_with("see", anchors_ws))
    print("v3 QA summary:", qa_summary)


if __name__ == "__main__":
    main()
